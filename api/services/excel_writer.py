"""
Excel writer service for storing extracted data
Creates and updates Excel files with structured data
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Optional
from datetime import datetime
import os
from pathlib import Path


class ExcelWriter:
    """Write extracted data to Excel files"""
    
    def __init__(self, excel_file_path: str = "extracted_content.xlsx"):
        # Convert to absolute path to ensure file is created in the right location
        if not os.path.isabs(excel_file_path):
            # Get the project root directory (parent of api folder)
            # __file__ is api/services/excel_writer.py, so:
            # - dirname(__file__) = api/services
            # - dirname(api/services) = api
            # - dirname(api) = project root
            current_file_dir = os.path.dirname(os.path.abspath(__file__))
            api_dir = os.path.dirname(current_file_dir)  # api/services -> api
            project_root = os.path.dirname(api_dir)  # api -> project root
            self.excel_file_path = os.path.join(project_root, excel_file_path)
        else:
            self.excel_file_path = excel_file_path
        self.workbook: Optional[Workbook] = None
        print(f"[ExcelWriter] Initialized with file path: {self.excel_file_path}")
    
    def initialize_workbook(self):
        """Create or load Excel workbook with single sheet for extracted content"""
        expected_headers = ['id', 'profile_name', 'profile_url', 'all_urls', 'title', 'raw_text', 'cleaned_text',
                           'raw_headings', 'raw_paragraphs', 'cleaning_status', 'cleaning_method', 
                           'scraped_at', 'cleaned_at', 'word_count_raw', 'word_count_cleaned']
        
        if os.path.exists(self.excel_file_path):
            try:
                self.workbook = openpyxl.load_workbook(self.excel_file_path)
                print(f"[ExcelWriter] Loaded existing workbook: {self.excel_file_path}")
                
                # Check if structure matches - if not, recreate
                if 'Extracted_Content' in self.workbook.sheetnames:
                    sheet = self.workbook['Extracted_Content']
                    if sheet.max_row > 0:
                        # Check header row
                        existing_headers = [cell.value for cell in sheet[1]]
                        if existing_headers != expected_headers:
                            print(f"[ExcelWriter] Column structure mismatch detected. Recreating file with new structure...")
                            # Close old file
                            self.workbook.close()
                            # Try to delete, if locked, rename it
                            try:
                                os.remove(self.excel_file_path)
                                print(f"[ExcelWriter] Deleted old file")
                            except PermissionError:
                                # File is locked, rename it
                                import datetime
                                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                                backup_path = self.excel_file_path.replace('.xlsx', f'_old_{timestamp}.xlsx')
                                try:
                                    os.rename(self.excel_file_path, backup_path)
                                    print(f"[ExcelWriter] Renamed locked file to: {os.path.basename(backup_path)}")
                                except:
                                    print(f"[ExcelWriter] Warning: Could not rename locked file. Creating new file anyway...")
                            # Create new workbook - use create_sheet instead of Workbook() to avoid default sheet
                            self.workbook = Workbook()
                            # CRITICAL: Remove default sheet BEFORE creating new one
                            if 'Sheet' in self.workbook.sheetnames:
                                default_sheet = self.workbook['Sheet']
                                self.workbook.remove(default_sheet)
                                print(f"[ExcelWriter] Removed default 'Sheet'")
                        else:
                            # Structure matches, ensure no extra sheets exist
                            self._create_sheets()  # This will remove any extra sheets
                            return
                else:
                    # Extracted_Content sheet doesn't exist, recreate
                    self.workbook.close()
                    try:
                        os.remove(self.excel_file_path)
                    except:
                        pass
                    self.workbook = Workbook()
                    if 'Sheet' in self.workbook.sheetnames:
                        self.workbook.remove(self.workbook['Sheet'])
            except Exception as e:
                error_msg = str(e)
                print(f"[ExcelWriter] Error loading existing workbook: {error_msg}. Creating new file...")
                # If file is corrupted or locked, try to rename it
                if 'Permission' in error_msg or 'locked' in error_msg.lower() or 'corrupt' in error_msg.lower():
                    try:
                        import datetime
                        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        backup_path = self.excel_file_path.replace('.xlsx', f'_corrupted_{timestamp}.xlsx')
                        if os.path.exists(self.excel_file_path):
                            try:
                                os.rename(self.excel_file_path, backup_path)
                                print(f"[ExcelWriter] Renamed corrupted/locked file to: {os.path.basename(backup_path)}")
                            except:
                                print(f"[ExcelWriter] Could not rename file (locked). Will create new file.")
                    except:
                        pass
                else:
                    # Try to delete if not locked
                    try:
                        if os.path.exists(self.excel_file_path):
                            os.remove(self.excel_file_path)
                    except:
                        pass
                # Create new workbook
                self.workbook = Workbook()
                # CRITICAL: Remove default sheet immediately
                if 'Sheet' in self.workbook.sheetnames:
                    default_sheet = self.workbook['Sheet']
                    self.workbook.remove(default_sheet)
                    print(f"[ExcelWriter] Removed default 'Sheet'")
        else:
            print(f"[ExcelWriter] Creating new workbook: {self.excel_file_path}")
            self.workbook = Workbook()
            # CRITICAL: Remove default sheet immediately before creating new one
            if 'Sheet' in self.workbook.sheetnames:
                default_sheet = self.workbook['Sheet']
                self.workbook.remove(default_sheet)
                print(f"[ExcelWriter] Removed default 'Sheet'")
        
        # Create single sheet if it doesn't exist (this also removes any extra sheets)
        self._create_sheets()
        self._format_headers()
        
        # Save immediately after creation to ensure file exists
        try:
            # Ensure directory exists
            file_dir = os.path.dirname(self.excel_file_path)
            if file_dir and not os.path.exists(file_dir):
                os.makedirs(file_dir, exist_ok=True)
            
            # CRITICAL: Verify no default sheet exists before saving
            if 'Sheet' in self.workbook.sheetnames:
                print(f"[ExcelWriter] WARNING: Default 'Sheet' still exists, removing before save...")
                self.workbook.remove(self.workbook['Sheet'])
            
            # Verify we have exactly one sheet
            if len(self.workbook.sheetnames) != 1 or 'Extracted_Content' not in self.workbook.sheetnames:
                print(f"[ExcelWriter] WARNING: Sheet structure incorrect. Sheets: {self.workbook.sheetnames}")
                # Fix it
                for sheet_name in self.workbook.sheetnames:
                    if sheet_name != 'Extracted_Content':
                        self.workbook.remove(self.workbook[sheet_name])
                if 'Extracted_Content' not in self.workbook.sheetnames:
                    self.workbook.create_sheet('Extracted_Content')
                    self._format_headers()
            
            # Try to save, if file is locked, use a temporary name
            try:
                self.workbook.save(self.excel_file_path)
                print(f"[ExcelWriter] Created and saved new Excel file: {self.excel_file_path}")
            except PermissionError:
                # File is locked, save with timestamp
                import datetime
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_path = self.excel_file_path.replace('.xlsx', f'_{timestamp}.xlsx')
                self.workbook.save(temp_path)
                print(f"[ExcelWriter] WARNING: Original file locked. Saved to: {os.path.basename(temp_path)}")
                print(f"[ExcelWriter] Please close Excel and delete the old file, then rename this one.")
        except Exception as e:
            print(f"[ExcelWriter] Warning: Could not save new workbook: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _create_sheets(self):
        """Create single sheet for extracted content"""
        sheet_name = 'Extracted_Content'
        
        # CRITICAL: Remove ALL sheets first, then create only the one we need
        # This ensures no leftover XML/formula references from default sheets
        sheets_to_remove = list(self.workbook.sheetnames)
        for sheet_name_to_remove in sheets_to_remove:
            try:
                sheet_to_remove = self.workbook[sheet_name_to_remove]
                self.workbook.remove(sheet_to_remove)
                print(f"[ExcelWriter] Removed sheet: {sheet_name_to_remove}")
            except Exception as e:
                print(f"[ExcelWriter] Warning: Could not remove sheet {sheet_name_to_remove}: {str(e)}")
        
        # Create Extracted_Content sheet (now the only sheet)
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
            print(f"[ExcelWriter] Created sheet: {sheet_name}")
        
        # Final verification - ensure only one sheet exists
        if len(self.workbook.sheetnames) != 1 or self.workbook.sheetnames[0] != sheet_name:
            print(f"[ExcelWriter] ERROR: Sheet structure incorrect after creation. Sheets: {self.workbook.sheetnames}")
            # Force fix - remove all and recreate
            for s in list(self.workbook.sheetnames):
                try:
                    self.workbook.remove(self.workbook[s])
                except:
                    pass
            self.workbook.create_sheet(sheet_name)
            print(f"[ExcelWriter] Force recreated sheet: {sheet_name}")
    
    def _format_headers(self):
        """Format headers for the single Extracted_Content sheet"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Extracted_Content headers - single sheet for all extracted content
        if 'Extracted_Content' in self.workbook.sheetnames:
            sheet = self.workbook['Extracted_Content']
            headers = [
                'id', 'profile_name', 'profile_url', 'all_urls', 'title', 'raw_text', 'cleaned_text',
                'raw_headings', 'raw_paragraphs', 'cleaning_status',
                'cleaning_method', 'scraped_at', 'cleaned_at', 'word_count_raw', 'word_count_cleaned'
            ]
            self._set_headers(sheet, headers, header_fill, header_font)
            
            # Set column widths for better readability
            sheet.column_dimensions['A'].width = 12  # id
            sheet.column_dimensions['B'].width = 30  # profile_name
            sheet.column_dimensions['C'].width = 50  # profile_url
            sheet.column_dimensions['D'].width = 80  # all_urls (comma-separated)
            sheet.column_dimensions['E'].width = 40  # title
            sheet.column_dimensions['F'].width = 80  # raw_text
            sheet.column_dimensions['G'].width = 80  # cleaned_text
            sheet.column_dimensions['H'].width = 60  # raw_headings
            sheet.column_dimensions['I'].width = 60  # raw_paragraphs
            sheet.column_dimensions['J'].width = 15  # cleaning_status
            sheet.column_dimensions['K'].width = 20  # cleaning_method
            sheet.column_dimensions['L'].width = 20  # scraped_at
            sheet.column_dimensions['M'].width = 20  # cleaned_at
            sheet.column_dimensions['N'].width = 15  # word_count_raw
            sheet.column_dimensions['O'].width = 15  # word_count_cleaned
    
    def _set_headers(self, sheet, headers, fill, font):
        """Set headers for a sheet"""
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def write_faculty_profile(self, profile_id: str, profile_data: Dict, metadata: Dict):
        """Write faculty profile - NOTE: This method is deprecated. Use write_raw_content instead."""
        # Skip if Faculty_Profiles sheet doesn't exist (we only use Extracted_Content now)
        if 'Faculty_Profiles' not in self.workbook.sheetnames:
            print("[ExcelWriter] Skipping write_faculty_profile - Faculty_Profiles sheet not available (using single sheet mode)")
            return
        sheet = self.workbook['Faculty_Profiles']
        
        row = [
            profile_id,
            profile_data.get('name', ''),
            profile_data.get('university', ''),
            profile_data.get('department', ''),
            profile_data.get('profile_url', ''),
            profile_data.get('email', ''),
            profile_data.get('position', ''),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            metadata.get('extraction_status', 'Complete'),
            metadata.get('total_resources_found', 0),
            metadata.get('resources_processed', 0),
            metadata.get('sources_used', '')
        ]
        
        sheet.append(row)
    
    def write_extracted_resource(self, profile_id: str, resource_data: Dict):
        """Write extracted resource - NOTE: This method is deprecated. Use write_raw_content instead."""
        # Skip if Extracted_Resources sheet doesn't exist (we only use Extracted_Content now)
        if 'Extracted_Resources' not in self.workbook.sheetnames:
            print("[ExcelWriter] Skipping write_extracted_resource - Extracted_Resources sheet not available (using single sheet mode)")
            return
        sheet = self.workbook['Extracted_Resources']
        
        row = [
            profile_id,
            resource_data.get('url', ''),
            resource_data.get('resource_type', ''),
            resource_data.get('category', ''),
            resource_data.get('link_text', ''),
            resource_data.get('file_type', ''),
            resource_data.get('word_count', 0),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        sheet.append(row)
    
    def write_merged_data(self, profile_id: str, merged_data: Dict):
        """Write merged data to Merged_Data sheet"""
        sheet = self.workbook['Merged_Data']
        
        # Calculate counts
        publications_count = len(merged_data.get('publications', []))
        awards_count = len(merged_data.get('awards', []))
        education_count = len(merged_data.get('education', []))
        
        # Calculate data quality score (0-100)
        score = self._calculate_quality_score(merged_data)
        
        row = [
            profile_id,
            merged_data.get('name', ''),
            merged_data.get('university', ''),
            merged_data.get('department', ''),
            merged_data.get('bio', '')[:500],  # Truncate long bios
            ', '.join(merged_data.get('expertise', []))[:200],
            publications_count,
            awards_count,
            education_count,
            merged_data.get('sources_used', ''),
            score
        ]
        
        sheet.append(row)
    
    def write_publications(self, profile_id: str, publications: List[Dict], source: str):
        """Write publications to Publications sheet"""
        sheet = self.workbook['Publications']
        
        for pub in publications:
            row = [
                profile_id,
                pub.get('type', ''),
                pub.get('title', ''),
                pub.get('authors', ''),
                pub.get('year', ''),
                pub.get('journal', '') or pub.get('publisher', ''),
                pub.get('doi', ''),
                pub.get('link', ''),
                source
            ]
            sheet.append(row)
    
    def write_education(self, profile_id: str, education: List[Dict], source: str):
        """Write education to Education sheet"""
        sheet = self.workbook['Education']
        
        for edu in education:
            # Handle both dict and string formats
            if isinstance(edu, dict):
                row = [
                    profile_id,
                    edu.get('degree', '') or edu.get('degree_type', ''),
                    edu.get('field', '') or edu.get('field_of_study', ''),
                    edu.get('institution', ''),
                    edu.get('year', ''),
                    edu.get('thesis', '') or edu.get('thesis_title', ''),
                    source
                ]
            else:
                # If it's just a string description
                row = [
                    profile_id,
                    '', '', edu, '', '', source
                ]
            sheet.append(row)
    
    def write_awards(self, profile_id: str, awards: List[Dict], source: str):
        """Write awards to Awards_Milestones sheet"""
        sheet = self.workbook['Awards_Milestones']
        
        for award in awards:
            if isinstance(award, dict):
                row = [
                    profile_id,
                    award.get('name', ''),
                    award.get('year', ''),
                    award.get('organization', ''),
                    award.get('description', ''),
                    source
                ]
            else:
                # If it's just a string
                row = [
                    profile_id,
                    award, '', '', '', source
                ]
            sheet.append(row)
    
    def write_expertise(self, profile_id: str, expertise: List, source: str):
        """Write expertise to Research_Expertise sheet"""
        sheet = self.workbook['Research_Expertise']
        
        for exp in expertise:
            if isinstance(exp, dict):
                row = [
                    profile_id,
                    exp.get('area', '') or exp.get('expertise_area', ''),
                    exp.get('description', ''),
                    source
                ]
            else:
                # If it's just a string
                row = [
                    profile_id,
                    exp, '', source
                ]
            sheet.append(row)
    
    def write_experience(self, profile_id: str, experience: List[Dict], source: str):
        """Write professional experience to Professional_Experience sheet"""
        sheet = self.workbook['Professional_Experience']
        
        for exp in experience:
            if isinstance(exp, dict):
                row = [
                    profile_id,
                    exp.get('position', '') or exp.get('position_title', ''),
                    exp.get('institution', ''),
                    exp.get('start_year', '') or exp.get('start_date', ''),
                    exp.get('end_year', '') or exp.get('end_date', '') or 'Current',
                    exp.get('description', ''),
                    source
                ]
            else:
                row = [
                    profile_id,
                    exp, '', '', '', '', source
                ]
            sheet.append(row)
    
    def write_profile_content(self, profile_id: str, profile_name: str, profile_url: str, 
                             all_urls: List[str], combined_text: str, 
                             combined_headings: List[str] = None, combined_paragraphs: List[str] = None,
                             cleaned_text: str = None, cleaning_status: str = 'pending',
                             cleaning_method: str = None) -> None:
        """
        Write all content for a single profile in ONE row
        
        Args:
            profile_id: Unique identifier for this profile
            profile_name: Name of the profile
            profile_url: Main profile URL
            all_urls: List of all URLs scraped (profile + documents + webpages)
            combined_text: Combined raw text from all sources
            combined_headings: Combined headings from all sources
            combined_paragraphs: Combined paragraphs from all sources
            cleaned_text: Cleaned text (optional, can be added later)
            cleaning_status: Status of cleaning (pending/cleaned/failed)
            cleaning_method: Method used for cleaning (regex/llm/manual)
        """
        # Ensure workbook is initialized
        if self.workbook is None:
            self.initialize_workbook()
        
        # Ensure Extracted_Content sheet exists
        if 'Extracted_Content' not in self.workbook.sheetnames:
            self._create_sheets()
            self._format_headers()
        
        sheet = self.workbook['Extracted_Content']
        
        # Convert URLs list to comma-separated string
        all_urls_str = ', '.join(all_urls) if all_urls else ''
        
        # Convert lists to JSON strings for storage
        headings_str = ''
        paragraphs_str = ''
        if combined_headings:
            import json
            headings_str = json.dumps(combined_headings, ensure_ascii=False)
        if combined_paragraphs:
            import json
            paragraphs_str = json.dumps(combined_paragraphs, ensure_ascii=False)
        
        # Calculate word counts
        word_count_raw = len(combined_text.split()) if combined_text else 0
        word_count_cleaned = len(cleaned_text.split()) if cleaned_text else 0
        
        # Get title (use profile name or first part of profile URL)
        title = profile_name or profile_url.split('/')[-1] or 'Profile'
        
        row = [
            profile_id,
            profile_name or '',
            profile_url,
            all_urls_str,  # All URLs comma-separated
            title,
            combined_text or '',  # Combined raw text from all sources
            cleaned_text or '',  # Empty initially, filled after cleaning
            headings_str,
            paragraphs_str,
            cleaning_status,
            cleaning_method or '',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # scraped_at
            '',  # cleaned_at (empty initially)
            word_count_raw,
            word_count_cleaned
        ]
        
        try:
            sheet.append(row)
            print(f"[ExcelWriter] SUCCESS: Wrote profile content - ID: {profile_id}, Name: {profile_name}, URLs: {len(all_urls)}, Text length: {len(combined_text)}")
            print(f"[ExcelWriter] Row written to sheet '{sheet.title}', Row number: {sheet.max_row}")
        except Exception as e:
            print(f"[ExcelWriter] ERROR: Failed to append row: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def write_raw_content(self, content_id: str, url: str, title: str, raw_text: str, 
                         raw_headings: List[str] = None, raw_paragraphs: List[str] = None,
                         cleaned_text: str = None, cleaning_status: str = 'pending',
                         cleaning_method: str = None) -> None:
        """
        DEPRECATED: Use write_profile_content instead for single row per profile
        Kept for backward compatibility - redirects to write_profile_content
        """
        # Redirect to new method for single row per profile
        self.write_profile_content(
            profile_id=content_id,
            profile_name=title,
            profile_url=url,
            all_urls=[url],
            combined_text=raw_text,
            combined_headings=raw_headings or [],
            combined_paragraphs=raw_paragraphs or [],
            cleaned_text=cleaned_text,
            cleaning_status=cleaning_status,
            cleaning_method=cleaning_method
        )
    
    def update_cleaned_content(self, content_id: str, cleaned_text: str, 
                               cleaning_method: str = 'regex', status: str = 'cleaned') -> bool:
        """
        Update cleaned text for existing raw content entry
        
        Args:
            content_id: ID of the content to update
            cleaned_text: Cleaned text content
            cleaning_method: Method used (regex/llm/manual)
            status: Cleaning status (cleaned/failed)
        
        Returns:
            True if updated successfully, False if not found
        """
        if 'Extracted_Content' not in self.workbook.sheetnames:
            return False
        
        sheet = self.workbook['Extracted_Content']
        
        # Find row with matching content_id (column A)
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == content_id:
                # Update cleaned_text (column G - new structure)
                sheet.cell(row=row_idx, column=7).value = cleaned_text
                # Update cleaning_status (column J - new structure)
                sheet.cell(row=row_idx, column=10).value = status
                # Update cleaning_method (column K - new structure)
                sheet.cell(row=row_idx, column=11).value = cleaning_method
                # Update cleaned_at (column M - new structure)
                sheet.cell(row=row_idx, column=13).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # Update word_count_cleaned (column O - new structure)
                word_count_cleaned = len(cleaned_text.split()) if cleaned_text else 0
                sheet.cell(row=row_idx, column=15).value = word_count_cleaned
                return True
        
        return False
    
    def get_pending_content(self) -> List[Dict]:
        """
        Get all content entries with cleaning_status = 'pending'
        
        Returns:
            List of dictionaries with content data
        """
        if 'Extracted_Content' not in self.workbook.sheetnames:
            return []
        
        sheet = self.workbook['Extracted_Content']
        pending_items = []
        
        # Headers: id, url, title, raw_text, cleaned_text, raw_headings, raw_paragraphs,
        #          cleaning_status, cleaning_method, scraped_at, cleaned_at, word_count_raw, word_count_cleaned
        
        # Headers: id, profile_name, profile_url, all_urls, title, raw_text, cleaned_text, raw_headings, raw_paragraphs,
        #          cleaning_status, cleaning_method, scraped_at, cleaned_at, word_count_raw, word_count_cleaned
        
        for row_idx in range(2, sheet.max_row + 1):
            status = sheet.cell(row=row_idx, column=10).value  # cleaning_status column (J)
            if status == 'pending':
                import json
                item = {
                    'id': sheet.cell(row=row_idx, column=1).value,
                    'url': sheet.cell(row=row_idx, column=3).value,  # profile_url
                    'title': sheet.cell(row=row_idx, column=5).value,  # title
                    'raw_text': sheet.cell(row=row_idx, column=6).value,  # raw_text
                    'raw_headings': json.loads(sheet.cell(row=row_idx, column=8).value or '[]'),  # raw_headings
                    'raw_paragraphs': json.loads(sheet.cell(row=row_idx, column=9).value or '[]'),  # raw_paragraphs
                }
                pending_items.append(item)
        
        return pending_items
    
    def _calculate_quality_score(self, data: Dict) -> int:
        """Calculate data quality score (0-100)"""
        score = 0
        
        # Bio (20 points)
        if data.get('bio') and len(data['bio']) > 100:
            score += 20
        elif data.get('bio'):
            score += 10
        
        # Publications (20 points)
        if len(data.get('publications', [])) > 10:
            score += 20
        elif len(data.get('publications', [])) > 5:
            score += 15
        elif len(data.get('publications', [])) > 0:
            score += 10
        
        # Education (20 points)
        if len(data.get('education', [])) > 0:
            score += 20
        
        # Expertise (20 points)
        if len(data.get('expertise', [])) > 5:
            score += 20
        elif len(data.get('expertise', [])) > 0:
            score += 15
        
        # Awards/Experience (20 points)
        if len(data.get('awards', [])) > 0 or len(data.get('experience', [])) > 0:
            score += 20
        
        return min(score, 100)
    
    def save(self):
        """Save workbook to file"""
        if self.workbook is None:
            print("[ExcelWriter] Warning: Workbook is None, cannot save. Initializing...")
            self.initialize_workbook()
        
        try:
            # CRITICAL: Before saving, ensure no default 'Sheet' exists
            if 'Sheet' in self.workbook.sheetnames:
                print(f"[ExcelWriter] WARNING: Default 'Sheet' found before save, removing...")
                try:
                    self.workbook.remove(self.workbook['Sheet'])
                except:
                    pass
            
            # Verify we have exactly one sheet named 'Extracted_Content'
            if len(self.workbook.sheetnames) != 1 or 'Extracted_Content' not in self.workbook.sheetnames:
                print(f"[ExcelWriter] WARNING: Sheet structure incorrect before save. Sheets: {self.workbook.sheetnames}")
                # Fix it
                for sheet_name in list(self.workbook.sheetnames):
                    if sheet_name != 'Extracted_Content':
                        try:
                            self.workbook.remove(self.workbook[sheet_name])
                        except:
                            pass
                if 'Extracted_Content' not in self.workbook.sheetnames:
                    self.workbook.create_sheet('Extracted_Content')
                    self._format_headers()
            
            # Ensure directory exists
            file_dir = os.path.dirname(self.excel_file_path)
            if file_dir and not os.path.exists(file_dir):
                os.makedirs(file_dir, exist_ok=True)
            
            self.workbook.save(self.excel_file_path)
            print(f"[ExcelWriter] Excel file saved: {self.excel_file_path}")
        except Exception as e:
            print(f"[ExcelWriter] Error saving Excel file to {self.excel_file_path}: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def close(self):
        """Close workbook"""
        if self.workbook:
            self.save()
            self.workbook.close()


# Singleton instance
_excel_writer = None

def get_excel_writer(excel_file_path: str = "extracted_content.xlsx") -> ExcelWriter:
    """Get or create Excel writer instance"""
    global _excel_writer
    
    # Convert to absolute path for comparison (same logic as ExcelWriter.__init__)
    if not os.path.isabs(excel_file_path):
        current_file_dir = os.path.dirname(os.path.abspath(__file__))
        api_dir = os.path.dirname(current_file_dir)  # api/services -> api
        project_root = os.path.dirname(api_dir)  # api -> project root
        abs_excel_file_path = os.path.join(project_root, excel_file_path)
    else:
        abs_excel_file_path = excel_file_path
    
    # Create new instance if None or if file path changed
    if _excel_writer is None or _excel_writer.excel_file_path != abs_excel_file_path:
        _excel_writer = ExcelWriter(excel_file_path)
        _excel_writer.initialize_workbook()
        # Ensure file is saved immediately after initialization
        if not os.path.exists(_excel_writer.excel_file_path):
            _excel_writer.save()
    
    return _excel_writer



