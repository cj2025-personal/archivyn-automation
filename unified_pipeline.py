"""
Unified Pipeline: Excel → Scraping → Cleaning → Chunking → JSON Output
Complete end-to-end pipeline that processes Excel files and creates chunked JSON files
"""
import os
import sys
import json
import uuid
import asyncio
import traceback
import pandas as pd
import re
from collections import Counter
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple, Set
from datetime import datetime, timezone
from tqdm import tqdm
from dotenv import load_dotenv
from urllib.parse import urlparse

# Load environment variables
load_dotenv()

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

# Import services
from api.services.scraper import get_scraper
from api.services.data_cleaning_service import get_data_cleaning_service
from api.services.document_processor import get_document_processor
from api.utils.llm_text_cleaner import get_llm_text_cleaner
from api.utils.source_registry import SourceRegistry, SourceRecord
from api.utils.source_guardrails import (
    normalize_url,
    make_source_id,
    compute_text_hash,
    compute_simhash,
    detect_language,
    detect_pii,
    compute_quality_metrics,
    extract_title_from_text,
    infer_license_for_url,
    is_noise_domain,
)
from api.utils.robots_checker import RobotsCache
from api.utils.claim_extractor import extract_claims
from profile_chunking_pipeline import ProfileChunkingPipeline


def load_urls_file(file_path: str) -> List[str]:
    """Load URLs from a text file (one URL per line).

    Skips comment lines and silently drops URLs whose host is on the
    curated noise blocklist (PBS support pages, scribd listing pages,
    pdfcoffee, social hosts, ...). This prevents an obviously-junk URL
    pasted into the curated list from being scraped at all — Van
    Sertima's prior run, for example, had 7 ``help.pbs.org`` URLs that
    silently became "biography" sources.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"URLs file not found: {file_path}")
    urls: List[str] = []
    skipped: List[str] = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("#"):
                continue
            if is_noise_domain(line):
                skipped.append(line)
                continue
            urls.append(line)
    if skipped:
        print(f"[load_urls_file] Skipped {len(skipped)} noise-domain URL(s): "
              f"{', '.join(skipped[:3])}{'...' if len(skipped) > 3 else ''}")
    return urls


def merge_url_lists(url_lists: List[List[str]], seed_urls: Optional[List[str]] = None) -> List[str]:
    """Merge URL lists while preserving order and removing duplicates."""
    seen = set()
    merged: List[str] = []

    def add_url(u: str) -> None:
        if not u:
            return
        if u in seen:
            return
        seen.add(u)
        merged.append(u)

    if seed_urls:
        for u in seed_urls:
            add_url(u)

    for url_list in url_lists:
        for u in url_list:
            add_url(u)

    return merged


class UnifiedPipeline:
    """
    Complete pipeline from Excel to chunked JSON files
    Steps:
    1. Read Excel file
    2. Scrape profile URLs
    3. Clean extracted text
    4. Create section-aware chunks
    5. Save to JSON files
    """
    
    def __init__(
        self,
        output_dir: str = "output",
        chunking_output_dir: str = "output/chunked_profiles",
        use_llm_chunking: bool = True,
        use_llm_cleaning: Optional[bool] = None,
        llm_provider: str = "openai",
        llm_model: str = "gpt-4o-mini",
        source_registry_path: Optional[str] = None,
        strict_source_policy: Optional[bool] = None,
        default_search_meta: Optional[Dict[str, Any]] = None,
        incremental_sync_enabled: Optional[bool] = None,
        incremental_sync_batch_size: int = 100,
        incremental_pinecone_batch_size: int = 50,
        incremental_skip_pinecone: bool = False,
        incremental_skip_mongo: bool = False,
        incremental_skip_indexes: bool = False,
    ):
        """
        Initialize the unified pipeline
        
        Args:
            output_dir: Base output directory for intermediate files
            chunking_output_dir: Directory for final chunked profiles
            use_llm_chunking: Use LLM-based section-aware chunking (default: True)
            use_llm_cleaning: Use LLM-based cleanup before chunking (default: auto)
            llm_provider: LLM provider for chunking ("ollama" or "openai")
            llm_model: LLM model name
        """
        self.output_dir = Path(output_dir)
        self.chunking_output_dir = Path(chunking_output_dir)
        self.use_llm_chunking = use_llm_chunking
        # Debug flag: when enabled, logs all scraped link categories/URLs
        self.debug_links = str(os.getenv("DEBUG_SCRAPER_LINKS", "")).lower() in ("1", "true", "yes", "on")
        self.error_log_path = self.output_dir / "errors.log"
        env_llm_cleaning = os.getenv("USE_LLM_CLEANING")
        if use_llm_cleaning is not None:
            self.use_llm_cleaning = bool(use_llm_cleaning)
        elif env_llm_cleaning is not None:
            self.use_llm_cleaning = str(env_llm_cleaning).lower() in ("1", "true", "yes", "on")
        else:
            # Auto-enable when LLM chunking is on and an API key is present.
            self.use_llm_cleaning = bool(self.use_llm_chunking and os.getenv("OPENAI_API_KEY"))
        self.llm_cleaner = get_llm_text_cleaner() if self.use_llm_cleaning else None
        self.strict_source_policy = (
            strict_source_policy if strict_source_policy is not None
            else str(os.getenv("STRICT_SOURCE_POLICY", "0")).lower() in ("1", "true", "yes", "on")
        )
        self.source_registry_path = source_registry_path or os.getenv("SOURCE_REGISTRY_PATH") or str(
            (self.output_dir / "source_registry.jsonl")
        )
        self.source_registry = SourceRegistry(self.source_registry_path)
        self.robots_cache = RobotsCache(user_agent=os.getenv("SCRAPER_USER_AGENT", "*"))
        self.default_search_meta = default_search_meta or {}
        # content-hash -> source_id index (for dedup)
        self._content_hash_index: Dict[str, str] = {}
        for rec in self.source_registry.iter_latest():
            if rec.content and isinstance(rec.content, dict):
                t_hash = rec.content.get("text_hash")
                if t_hash and t_hash not in self._content_hash_index:
                    self._content_hash_index[t_hash] = rec.source_id
        self.nav_terms_re = re.compile(
            r"\b(home|about|contact|search|menu|navigation|skip to content|privacy|terms|"
            r"copyright|all rights reserved|sitemap|back to top)\b",
            re.I,
        )
        self.page_counter_re = re.compile(r"(?i)\bpage\s+\d+\s+of\s+\d+\b")
        self.reference_id_re = re.compile(r"(?i)\breference id\b[:\s-]*[a-z0-9-]{6,}")
        self.compact_page_line_re = re.compile(r"(?i)^(?:\s*(?:page\s*)?\d+\s*(?:of|/)\s*\d+\s*){1,}$")
        self.anti_bot_line_re = re.compile(
            r"(?i)(access to this page has been denied|verify you are human|captcha|cloudflare|"
            r"attention required|proof of work|press\s*&\s*hold|press and hold|"
            r"confirm you are a human|checking your browser|not a bot|security check)"
        )
        allowed_intents = {
            "speech",
            "interview",
            "paper",
            "testimony",
            "policy_statement",
            "biography",
        }
        env_required = os.getenv("INTENT_REQUIRED_TYPES", "speech,interview,paper,testimony,policy_statement")
        requested_intents = [t.strip().lower() for t in env_required.split(",") if t.strip()]
        self.intent_required_types = [t for t in requested_intents if t in allowed_intents]
        if not self.intent_required_types:
            self.intent_required_types = ["speech", "interview", "paper", "testimony", "policy_statement"]
        self.intent_required_set = set(self.intent_required_types)
        self.intent_gating_enabled = (
            str(os.getenv("INTENT_GATING_ENABLED", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.intent_min_covered_types = max(1, int(os.getenv("INTENT_MIN_COVERED_TYPES", "2")))
        self.intent_min_sources = max(1, int(os.getenv("INTENT_MIN_SOURCES", "3")))
        self.intent_min_chunks = max(1, int(os.getenv("INTENT_MIN_CHUNKS", "6")))
        self.intent_min_source_confidence = float(os.getenv("INTENT_MIN_SOURCE_CONFIDENCE", "0.35"))
        self.intent_min_source_confidence = min(1.0, max(0.0, self.intent_min_source_confidence))
        self.intent_scan_text_chars = max(800, int(os.getenv("INTENT_SCAN_TEXT_CHARS", "6000")))
        self.intent_strict_domain_filter = (
            str(os.getenv("INTENT_STRICT_DOMAIN_FILTER", "0")).lower() in ("1", "true", "yes", "on")
        )
        self.intent_high_value_domain_hints = (
            ".gov",
            ".edu",
            "nber.org",
            "aeaweb.org",
            "academic.oup.com",
            "wiley.com",
            "springer.com",
            "cambridge.org",
            "jstor.org",
            "sciencedirect.com",
            "nature.com",
            "science.org",
            "mit.edu",
            "rpi.edu",
            "duke.edu",
            "nrc.gov",
            "congress.gov",
            "govinfo.gov",
            "house.gov",
            "senate.gov",
            "loc.gov",
            "npr.org",
            "nytimes.com",
            "theguardian.com",
        )
        self.intent_low_value_domains = {
            "grokipedia.com",
            "custom-powder.com",
            "aaespeakers.com",
            "reddit.com",
            "x.com",
            "twitter.com",
            "facebook.com",
            "instagram.com",
            "linkedin.com",
            "tiktok.com",
            "pinterest.com",
        }
        self.intent_patterns = {
            "speech": [
                re.compile(r"(?i)\b(speech|remarks|keynote|lecture|address|talk|commencement|fireside)\b"),
                re.compile(r"(?i)\b(prepared remarks|public address|opening remarks)\b"),
            ],
            "interview": [
                re.compile(r"(?i)\b(interview|q&a|q and a|conversation|podcast|oral history)\b"),
                re.compile(r"(?i)\b(fireside chat|sit down with|in conversation with)\b"),
            ],
            "paper": [
                re.compile(r"(?i)\b(working paper|journal|article|doi|abstract|methodology|findings)\b"),
                re.compile(r"(?i)\b(nber|repec|econpapers|ssrn|arxiv)\b"),
            ],
            "testimony": [
                re.compile(r"(?i)\b(testimony|hearing|committee|congressional|statement for the record)\b"),
                re.compile(r"(?i)\b(chrg-|hhrg-|senate hearing|house hearing)\b"),
            ],
            "policy_statement": [
                re.compile(r"(?i)\b(policy statement|position paper|white paper|recommendation|policy report)\b"),
                re.compile(r"(?i)\b(task force|advisory committee|commission report|public policy)\b"),
            ],
            "biography": [
                re.compile(r"(?i)\b(biography|bio|about|profile|awards?|honors?)\b"),
            ],
        }
        self.intent_source_type_hints = {
            "speech": {"speech", "remarks", "transcript"},
            "interview": {"interview", "podcast", "oral_history"},
            "paper": {"publication", "paper", "journal"},
            "testimony": {"testimony", "hearing"},
            "policy_statement": {"policy", "report"},
            "biography": {"profile_page", "biography"},
        }
        self.profile_role_filter_enabled = (
            str(os.getenv("PROFILE_ROLE_FILTER_ENABLED", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.profile_role_filter_use_llm = (
            str(os.getenv("PROFILE_ROLE_FILTER_USE_LLM", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.profile_role_filter_model = os.getenv("PROFILE_ROLE_FILTER_MODEL", "gpt-4o-mini")
        self.profile_role_filter_timeout = max(5, int(os.getenv("PROFILE_ROLE_FILTER_TIMEOUT", "20")))
        self.profile_role_filter_scan_chars = max(800, int(os.getenv("PROFILE_ROLE_FILTER_SCAN_CHARS", "5000")))
        self.source_quality_filter_enabled = (
            str(os.getenv("SOURCE_QUALITY_FILTER_ENABLED", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.profile_relevance_filter_enabled = (
            str(os.getenv("PROFILE_RELEVANCE_FILTER_ENABLED", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.profile_text_filter_enabled = (
            str(os.getenv("PROFILE_TEXT_FILTER_ENABLED", "1")).lower() in ("1", "true", "yes", "on")
        )
        self.role_include_re = re.compile(
            r"(?i)\b(?:assistant|associate|adjunct|clinical|visiting|research|distinguished|emeritus|teaching|full|endowed)?\s*professor(?:\s+of\s+practice)?\b|\bscholar\b"
        )
        self.role_exclude_re = re.compile(
            r"(?i)\b(staff|student|undergraduate|graduate|phd candidate|doctoral candidate|postdoc|postdoctoral|administrator|coordinator|manager|specialist|technician)\b"
        )
        self.role_context_hint_re = re.compile(
            r"(?i)\b(professor|scholar|faculty|staff|student|lecturer|instructor|researcher|postdoc|fellow|dean|chair|adjunct|emeritus)\b"
        )
        self.profile_role_filter_client = self._init_profile_role_filter_client()
        env_incremental_sync = os.getenv("INCREMENTAL_SYNC_ENABLED")
        if incremental_sync_enabled is not None:
            self.incremental_sync_enabled = bool(incremental_sync_enabled)
        elif env_incremental_sync is not None:
            self.incremental_sync_enabled = str(env_incremental_sync).lower() in ("1", "true", "yes", "on")
        else:
            self.incremental_sync_enabled = False
        self.incremental_sync_batch_size = max(1, int(incremental_sync_batch_size))
        self.incremental_pinecone_batch_size = max(1, int(incremental_pinecone_batch_size))
        self.incremental_skip_pinecone = bool(incremental_skip_pinecone)
        self.incremental_skip_mongo = bool(incremental_skip_mongo)
        self.incremental_skip_indexes = bool(incremental_skip_indexes)
        self._incremental_vector_db = None
        self._incremental_embeddings_service = None
        self._incremental_index_dimension: Optional[int] = None
        self._incremental_mongo_sync = None
        self._incremental_indexes_created = False
        self._incremental_sync_stats = {
            "enabled": bool(self.incremental_sync_enabled),
            "batch_size_profiles": self.incremental_sync_batch_size,
            "pinecone_batch_size_chunks": self.incremental_pinecone_batch_size,
            "skip_pinecone": bool(self.incremental_skip_pinecone),
            "skip_mongo": bool(self.incremental_skip_mongo),
            "skip_indexes": bool(self.incremental_skip_indexes),
            "batches_attempted": 0,
            "batches_completed": 0,
            "profiles_enqueued": 0,
            "profiles_synced_mongo": 0,
            "profiles_failed_mongo": 0,
            "chunks_loaded": 0,
            "vectors_uploaded": 0,
            "vectors_failed": 0,
            "last_error": "",
        }
        
        # Create output directories
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.chunking_output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize services
        print("[Pipeline] Initializing services...")
        self.scraper = None  # Will be initialized async
        self.cleaning_service = get_data_cleaning_service(
            target_words_per_chunk=325,
            min_words_per_chunk=250,
            max_words_per_chunk=400,
            use_llm_cleaning=self.use_llm_cleaning,
            llm_provider=(os.getenv("CLEANING_LLM_PROVIDER") or "openai"),
            llm_model=(os.getenv("CLEANING_LLM_MODEL") or "gpt-4o-mini"),
        )
        self.document_processor = get_document_processor()  # For CV extraction
        
        if use_llm_chunking:
            self.chunking_pipeline = ProfileChunkingPipeline(
                output_dir=str(chunking_output_dir),
                llm_provider=llm_provider,
                llm_model="gpt-4o-mini"
            )
        else:
            self.chunking_pipeline = None
        
        print(f"[Pipeline] LLM Cleaning: {'Enabled' if self.use_llm_cleaning else 'Disabled'}")
        llm_status = "Enabled" if self.profile_role_filter_client is not None else "FallbackOnly"
        print(
            "[Pipeline] Profile Role Filter: "
            f"{'Enabled' if self.profile_role_filter_enabled else 'Disabled'} | "
            f"LLM={llm_status}"
        )
        print(
            "[Pipeline] Source Filters: "
            f"quality={'On' if self.source_quality_filter_enabled else 'Off'}, "
            f"profile_relevance={'On' if self.profile_relevance_filter_enabled else 'Off'}, "
            f"profile_text={'On' if self.profile_text_filter_enabled else 'Off'}"
        )
        if self.incremental_sync_enabled:
            print(
                "[Pipeline] Incremental Sync: Enabled | "
                f"batch_profiles={self.incremental_sync_batch_size}, "
                f"pinecone_chunk_batch={self.incremental_pinecone_batch_size}, "
                f"pinecone={'Skip' if self.incremental_skip_pinecone else 'On'}, "
                f"mongo={'Skip' if self.incremental_skip_mongo else 'On'}"
            )
        else:
            print("[Pipeline] Incremental Sync: Disabled")
        print(
            "[Pipeline] Intent Gates: "
            f"{'Enabled' if self.intent_gating_enabled else 'Disabled'} | "
            f"required={','.join(self.intent_required_types)} | "
            f"min_types={self.intent_min_covered_types}, min_sources={self.intent_min_sources}, "
            f"min_chunks={self.intent_min_chunks}, min_conf={self.intent_min_source_confidence:.2f}"
        )
        print("[Pipeline] Services initialized")

    def _log_error(self, message: str) -> None:
        """Append an error message to the run error log."""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().isoformat()
            with open(self.error_log_path, "a", encoding="utf-8") as f:
                f.write(f"[{timestamp}] {message}\n")
        except Exception:
            # Last resort: avoid crashing on logging failures
            pass

    def _is_reasonable_text(self, text: str) -> bool:
        """Reject mostly-binary / garbled text blobs."""
        if not text:
            return False
        length = len(text)
        if length < 50:
            return False
        if length > 200000:  # skip oversized blobs
            return False
        ascii_chars = sum(1 for ch in text if 32 <= ord(ch) < 127)
        ratio = ascii_chars / max(1, length)
        return ratio >= 0.6

    @staticmethod
    def _root_domain_for_url(url: str) -> str:
        try:
            host = (urlparse(url).netloc or "").lower()
        except Exception:
            return ""
        if host.startswith("www."):
            host = host[4:]
        parts = [part for part in host.split(".") if part]
        if len(parts) >= 2:
            return ".".join(parts[-2:])
        return host

    def _allow_official_profile_page_despite_low_quality(
        self,
        *,
        content_text: str,
        source_type: str,
        source_url: str,
        profile_url: str,
        profile_name: str,
    ) -> bool:
        if source_type not in {"profile_page", "personal_website", "personal_website_subpage"}:
            return False
        if len((content_text or "").strip()) < 120:
            return False
        if not self._content_mentions_subject(content_text, profile_name):
            return False
        source_root = self._root_domain_for_url(source_url)
        profile_root = self._root_domain_for_url(profile_url)
        if not source_root or not profile_root:
            return False
        return source_root == profile_root

    def _is_relevant_to_profile(self, text: str, profile_name: str) -> bool:
        """Keep only text that strongly references the target person."""
        if not text:
            return False
        if not profile_name:
            return True
        name = self._normalize_inline_whitespace(profile_name).lower()
        if not name:
            return True
        text_norm = self._normalize_inline_whitespace(text).lower()
        if name in text_norm:
            return True
        no_period_name = name.replace(".", " ")
        no_period_name = re.sub(r"\s+", " ", no_period_name).strip()
        if no_period_name and no_period_name in text_norm:
            return True
        parts = [p for p in re.findall(r"[a-z0-9]+", name) if len(p) >= 2]
        if len(parts) >= 2:
            first = parts[0]
            last = parts[-1]
            if re.search(rf"\b{re.escape(first)}\b.*\b{re.escape(last)}\b", text_norm):
                return True
            initials = " ".join([p[:1] for p in parts[:-1]] + [last]).strip()
            if initials and initials in text_norm:
                return True
        return False

    @staticmethod
    def _document_like_url(url: str) -> bool:
        lower = (url or "").lower()
        return lower.endswith((".pdf", ".doc", ".docx")) or "/cv/" in lower or "/resume" in lower or "/vita" in lower

    @staticmethod
    def _generic_listing_url(url: str) -> bool:
        lower = (url or "").lower()
        tokens = [tok for tok in re.findall(r"[a-z0-9]+", lower) if tok]
        generic = {
            "team",
            "our",
            "index",
            "directory",
            "directories",
            "faculty",
            "staff",
            "people",
            "profiles",
            "profile",
            "former",
            "emeritus",
            "news",
        }
        return any(tok in generic for tok in tokens)

    @staticmethod
    def _name_identity_parts(profile_name: str) -> Tuple[str, str, str, List[str]]:
        normalized = re.sub(r"\s+", " ", (profile_name or "").strip()).lower()
        normalized = normalized.replace(".", " ")
        normalized = re.sub(r"\s+", " ", normalized).strip()
        parts = [tok for tok in re.findall(r"[a-z0-9]+", normalized) if tok]
        first = parts[0] if parts else ""
        last = parts[-1] if len(parts) >= 2 else ""
        variants: List[str] = []
        if normalized:
            variants.append(normalized)
        if parts:
            compact = " ".join(parts)
            if compact and compact not in variants:
                variants.append(compact)
        if first and last:
            first_last = f"{first} {last}"
            if first_last not in variants:
                variants.append(first_last)
            initials = " ".join([p[:1] for p in parts[:-1]] + [last]).strip()
            if initials and initials not in variants:
                variants.append(initials)
        return normalized, first, last, variants

    def _text_conflicts_with_profile_identity(self, text: str, profile_name: str) -> bool:
        if not text or not profile_name:
            return False
        _, first, last, _ = self._name_identity_parts(profile_name)
        if not first or not last:
            return False
        tokens = [tok for tok in re.findall(r"[a-z]+", text.lower()) if tok]
        if not tokens:
            return False
        first_initial = first[:1]
        ignore = {
            "dr",
            "mr",
            "mrs",
            "ms",
            "prof",
            "professor",
            "assistant",
            "associate",
            "adjunct",
            "emeritus",
            "former",
            "faculty",
            "deceased",
            "department",
            "program",
            "school",
            "college",
            "university",
            "of",
            "the",
            "and",
            "cv",
            "resume",
            "vitae",
            "pdf",
            "doc",
            "docx",
            "page",
            "profile",
            "team",
            "people",
            "person",
            "faculty",
            "staff",
            "directory",
            "news",
            "group",
            "lab",
            "research",
            "about",
        }
        for idx, tok in enumerate(tokens):
            if tok != last:
                continue
            neighbors: List[str] = []
            for offset in (-2, -1, 1, 2):
                probe = idx + offset
                if 0 <= probe < len(tokens):
                    neighbors.append(tokens[probe])
            matched_target = any(
                neighbor == first or (len(neighbor) <= 2 and neighbor[:1] == first_initial)
                for neighbor in neighbors
            )
            if matched_target:
                continue
            for neighbor in neighbors:
                if len(neighbor) < 3 or neighbor in ignore:
                    continue
                return True
        return False

    def _url_or_title_matches_profile_identity(self, text: str, profile_name: str) -> bool:
        if not text or not profile_name:
            return False
        lowered = re.sub(r"[_/\-]+", " ", text.lower())
        lowered = re.sub(r"\s+", " ", lowered).strip()
        _, first, last, variants = self._name_identity_parts(profile_name)
        compact = re.sub(r"[^a-z0-9]+", "", lowered)
        for variant in variants:
            variant_compact = re.sub(r"[^a-z0-9]+", "", variant)
            if variant and variant in lowered:
                return True
            if variant_compact and variant_compact in compact:
                return True
        return bool(first and last and first in lowered and last in lowered)

    def _document_front_matter_matches_profile(self, text: str, profile_name: str) -> bool:
        if not text or not profile_name:
            return False
        lowered = self._normalize_inline_whitespace(text).lower()
        _, first, last, variants = self._name_identity_parts(profile_name)
        probes = list(variants)
        positions: List[int] = []
        if first and last:
            regex = re.compile(
                rf"\b{re.escape(first)}\b(?:\s+[a-z]\.?)?(?:\s+[a-z]+)?\s+\b{re.escape(last)}\b"
            )
            match = regex.search(lowered)
            if match:
                positions.append(match.start())
        for probe in probes:
            if not probe:
                continue
            idx = lowered.find(probe)
            if idx >= 0:
                positions.append(idx)
        if not positions:
            return False
        return min(positions) <= 30

    def _source_matches_target_profile(
        self,
        *,
        source_type: str,
        source_url: str,
        link_text: str,
        content_text: str,
        profile_name: str,
    ) -> bool:
        if not profile_name:
            return True
        if not content_text:
            return False

        label = self._normalize_inline_whitespace(link_text)
        url = source_url or ""
        leading_text = content_text[:4000]
        subject_in_lead = self._is_relevant_to_profile(leading_text, profile_name)
        subject_in_full = self._content_mentions_subject(content_text, profile_name)
        title_matches = self._url_or_title_matches_profile_identity(label, profile_name)
        url_matches = self._url_or_title_matches_profile_identity(url, profile_name)
        title_conflicts = self._text_conflicts_with_profile_identity(label, profile_name)
        url_conflicts = self._text_conflicts_with_profile_identity(url, profile_name)

        if title_conflicts or url_conflicts:
            return False

        if source_type == "cv" or self._document_like_url(url):
            if not subject_in_lead:
                return False
            return (
                url_matches
                or title_matches
                or self._document_front_matter_matches_profile(leading_text[:1400], profile_name)
            )

        if self._generic_listing_url(url):
            if not subject_in_lead:
                return False
            if title_matches or url_matches:
                return True
            _, _, last, _ = self._name_identity_parts(profile_name)
            if not last:
                return False
            mention_count = content_text.lower().count(last)
            return mention_count >= 2

        return subject_in_full

    def _content_mentions_subject(self, text: str, profile_name: str) -> bool:
        """Page-level subject filter — drop pages clearly not about the subject.

        Used at the source-ingestion gate, *before* line-level cleaning. Any
        of these counts as "this page is about the subject":
          - the full normalised profile name appears at least once
          - the *first + last* names co-occur
          - the last name appears at least twice (a single drive-by mention
            in a list of names doesn't count)
        Plus a *density* check for long pages: the subject must be
        mentioned at least once per ~2,500 words, otherwise the page is
        treated as a passing reference (anthology chapter, museum index)
        rather than biographical content.
        """
        if not text or not profile_name:
            return True
        text_lc = text.lower()
        name = profile_name.strip().lower()
        parts = [p for p in name.split() if len(p) >= 3]
        if not parts:
            return True
        first = parts[0]
        last = parts[-1]

        last_occurrences = text_lc.count(last) if last else 0
        full_name_present = bool(name and name in text_lc)
        first_present = bool(first and first in text_lc)
        initials = ".".join(p[0] for p in parts[:-1]) + "." if len(parts) > 1 else ""
        initials_present = bool(initials and initials.lower() in text_lc)

        passes_basic = (
            full_name_present
            or last_occurrences >= 2
            or (last_occurrences >= 1 and first_present)
            or (initials_present and last_occurrences >= 1)
        )
        if not passes_basic:
            return False

        # Density check: a 30,000-word anthology that name-checks the
        # subject once at the beginning is not "about" them.
        word_count = len(re.findall(r"\w+", text_lc))
        if word_count > 4000:
            min_required = max(2, word_count // 2500)
            if last_occurrences < min_required and not full_name_present:
                return False
        return True

    def _init_profile_role_filter_client(self):
        if not self.profile_role_filter_enabled:
            return None
        if not self.profile_role_filter_use_llm:
            return None
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return None
        try:
            from openai import OpenAI
            try:
                import httpx
                http_client = httpx.Client(timeout=float(self.profile_role_filter_timeout + 10))
                return OpenAI(api_key=api_key, http_client=http_client)
            except Exception:
                return OpenAI(api_key=api_key)
        except Exception as e:
            self._log_error(f"[profile_role_filter_client] init failed :: {e}")
            return None

    @staticmethod
    def _normalize_inline_whitespace(value: str) -> str:
        return re.sub(r"\s+", " ", (value or "")).strip()

    def _collect_profile_role_context(
        self,
        *,
        profile_name: str,
        profile_url: str,
        profile_data: Optional[Dict[str, Any]],
        row_data: Optional[Dict[str, Any]],
        combined_text: str,
    ) -> Dict[str, Any]:
        title_candidates: List[str] = []

        if profile_data and isinstance(profile_data, dict):
            for key in ("position", "title", "job_title", "rank", "designation"):
                value = self._normalize_inline_whitespace(str(profile_data.get(key, "")))
                if value:
                    title_candidates.append(value)

        if row_data and isinstance(row_data, dict):
            for key, raw_value in row_data.items():
                key_str = str(key)
                if key_str.startswith("_"):
                    continue
                key_lc = key_str.lower()
                if not any(tok in key_lc for tok in ("title", "position", "role", "rank", "designation", "job")):
                    continue
                value = self._normalize_inline_whitespace(str(raw_value))
                if not value:
                    continue
                if value.lower() in ("nan", "none", "null"):
                    continue
                title_candidates.append(value)

        deduped_titles: List[str] = []
        seen_titles = set()
        for title in title_candidates:
            key = title.lower()
            if key in seen_titles:
                continue
            seen_titles.add(key)
            deduped_titles.append(title)

        snippet_lines: List[str] = []
        scan_text = (combined_text or "")[: self.profile_role_filter_scan_chars]
        for raw_line in scan_text.splitlines():
            line = self._normalize_inline_whitespace(raw_line)
            if len(line) < 4:
                continue
            if not self.role_context_hint_re.search(line):
                continue
            snippet_lines.append(line[:260])
            if len(snippet_lines) >= 25:
                break

        if not snippet_lines and scan_text.strip():
            snippet_lines.append(self._normalize_inline_whitespace(scan_text[:300]))

        return {
            "profile_name": self._normalize_inline_whitespace(profile_name),
            "profile_url": profile_url or "",
            "title_candidates": deduped_titles[:10],
            "snippet_lines": snippet_lines,
        }

    def _heuristic_profile_role_decision(self, role_context: Dict[str, Any]) -> Dict[str, Any]:
        context_parts: List[str] = []
        context_parts.extend(role_context.get("title_candidates") or [])
        context_parts.extend(role_context.get("snippet_lines") or [])
        context_blob = " | ".join(context_parts)

        include_match = self.role_include_re.search(context_blob)
        if include_match:
            matched = self._normalize_inline_whitespace(include_match.group(0))
            return {
                "include": True,
                "reason": f"matched_role:{matched}",
                "matched_title": matched,
                "method": "heuristic",
            }

        exclude_match = self.role_exclude_re.search(context_blob)
        if exclude_match:
            label = self._normalize_inline_whitespace(exclude_match.group(0))
            return {
                "include": False,
                "reason": f"non_target_role:{label}",
                "matched_title": "",
                "method": "heuristic",
            }

        return {
            "include": False,
            "reason": "no_professor_or_scholar_title_detected",
            "matched_title": "",
            "method": "heuristic",
        }

    def _llm_profile_role_decision(self, role_context: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if self.profile_role_filter_client is None:
            return None

        prompt_payload = {
            "profile_name": role_context.get("profile_name", ""),
            "profile_url": role_context.get("profile_url", ""),
            "title_candidates": role_context.get("title_candidates", []),
            "role_snippets": role_context.get("snippet_lines", []),
        }

        try:
            response = self.profile_role_filter_client.chat.completions.create(
                model=self.profile_role_filter_model,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Classify whether an academic profile should be INCLUDED for professor/scholar ingestion.\n"
                            "Include only if the person appears to hold professor/scholar roles, e.g. professor, assistant professor, "
                            "associate professor, adjunct professor, visiting professor, clinical professor, research professor, "
                            "professor emeritus, professor of practice, or scholar.\n"
                            "Ignore staff, students, administrators, coordinators, and other non-professor profiles.\n"
                            "If uncertain, choose ignore.\n"
                            "Return ONLY JSON with keys: decision (include|ignore), matched_title, reason, confidence."
                        ),
                    },
                    {
                        "role": "user",
                        "content": json.dumps(prompt_payload, ensure_ascii=False),
                    },
                ],
                temperature=0.0,
                max_tokens=180,
                timeout=self.profile_role_filter_timeout,
                response_format={"type": "json_object"},
            )
        except Exception as e:
            self._log_error(f"[profile_role_filter_llm] request failed :: {e}")
            return None

        try:
            raw = (response.choices[0].message.content or "").strip()
            parsed = json.loads(raw) if raw else {}
        except Exception as e:
            self._log_error(f"[profile_role_filter_llm] parse failed :: {e}")
            return None

        decision = str(parsed.get("decision", "")).strip().lower()
        if decision not in ("include", "ignore"):
            return None

        matched_title = self._normalize_inline_whitespace(str(parsed.get("matched_title", "")))
        reason = self._normalize_inline_whitespace(str(parsed.get("reason", "")))
        confidence = parsed.get("confidence")
        confidence_value = None
        try:
            if confidence is not None:
                confidence_value = round(max(0.0, min(1.0, float(confidence))), 4)
        except Exception:
            confidence_value = None

        result = {
            "include": decision == "include",
            "reason": reason or ("llm_include" if decision == "include" else "llm_ignore"),
            "matched_title": matched_title,
            "method": "llm",
        }
        if confidence_value is not None:
            result["confidence"] = confidence_value
        return result

    def _should_include_profile_by_role(
        self,
        *,
        profile_name: str,
        profile_url: str,
        profile_data: Optional[Dict[str, Any]],
        row_data: Optional[Dict[str, Any]],
        combined_text: str,
    ) -> Dict[str, Any]:
        if not self.profile_role_filter_enabled:
            return {
                "include": True,
                "reason": "profile_role_filter_disabled",
                "matched_title": "",
                "method": "disabled",
            }

        role_context = self._collect_profile_role_context(
            profile_name=profile_name,
            profile_url=profile_url,
            profile_data=profile_data,
            row_data=row_data,
            combined_text=combined_text,
        )
        heuristic = self._heuristic_profile_role_decision(role_context)
        llm_result = self._llm_profile_role_decision(role_context)
        if llm_result is not None:
            llm_result["fallback_method"] = heuristic.get("method")
            llm_result["fallback_include"] = bool(heuristic.get("include"))
            return llm_result
        return heuristic

    def _filter_text_for_profile(self, text: str, profile_name: str) -> str:
        """
        Remove boilerplate/nav lines and very short lines unrelated to the profile.
        Keeps section markers and lines that mention the person's name.
        """
        if not self.profile_text_filter_enabled:
            return text or ""
        if not text:
            return ""
        name = (profile_name or "").strip().lower()
        parts = [p for p in name.split() if len(p) >= 3]
        last_name = parts[-1] if parts else ""

        filtered_lines = []
        seen_lines = set()
        for line in text.splitlines():
            raw = line.strip()
            if not raw:
                continue
            # Drop section markers like "=== SEED URL ===" or "=== PROFILE PAGE ==="
            if raw.startswith("===") and raw.endswith("==="):
                continue
            if raw.startswith("=== SEED URL") or raw.startswith("=== PROFILE PAGE"):
                continue
            # Drop anti-bot/blocked-access challenge pages and reference IDs
            if self.anti_bot_line_re.search(raw):
                continue
            if self.reference_id_re.search(raw):
                continue
            # Drop obvious pagination/OCR index noise lines.
            page_hits = len(self.page_counter_re.findall(raw))
            if page_hits >= 2 or self.compact_page_line_re.match(raw):
                continue
            # Drop boilerplate/nav lines
            if self.nav_terms_re.search(raw):
                continue
            words = raw.split()
            has_name = False
            raw_lc = raw.lower()
            if name and name in raw_lc:
                has_name = True
            elif last_name and last_name in raw_lc:
                has_name = True
            # Drop number-heavy lines unless they mention the person name.
            alpha_count = sum(1 for ch in raw if ch.isalpha())
            digit_count = sum(1 for ch in raw if ch.isdigit())
            if not has_name and digit_count >= 10 and alpha_count <= digit_count:
                continue
            # Drop very short lines unless they mention the person
            if len(words) < 5 and not has_name:
                continue
            line_key = re.sub(r"\s+", " ", raw_lc).strip()
            if line_key in seen_lines:
                continue
            seen_lines.add(line_key)
            filtered_lines.append(raw)

        return "\n".join(filtered_lines)

    def _build_raw_token_index(self, raw_text: str) -> Optional[Dict[str, Any]]:
        if not raw_text:
            return None
        tokens = []
        spans = []
        for match in re.finditer(r"[A-Za-z0-9]+", raw_text):
            tokens.append(match.group(0).lower())
            spans.append((match.start(), match.end()))
        if not tokens:
            return None
        return {"text": raw_text, "tokens": tokens, "spans": spans}

    def _find_token_sequence(self, tokens: List[str], sequence: List[str], start_idx: int) -> Optional[int]:
        if not sequence:
            return None
        max_start = len(tokens) - len(sequence)
        first = sequence[0]
        for i in range(start_idx, max_start + 1):
            if tokens[i] != first:
                continue
            if tokens[i:i + len(sequence)] == sequence:
                return i
        return None

    def _map_chunk_offsets(
        self,
        raw_index: Optional[Dict[str, Any]],
        chunk_text: str,
        start_token_idx: int
    ) -> Tuple[int, int, int]:
        """
        Best-effort mapping of cleaned chunk text back to raw text offsets.
        Returns (raw_start, raw_end, next_start_token_idx).
        """
        if not raw_index or not chunk_text:
            return -1, -1, start_token_idx

        raw_tokens = raw_index["tokens"]
        raw_spans = raw_index["spans"]
        chunk_tokens = [m.group(0).lower() for m in re.finditer(r"[A-Za-z0-9]+", chunk_text)]
        if not chunk_tokens:
            return -1, -1, start_token_idx

        anchor_len = min(6, len(chunk_tokens))
        anchor_pos = None
        for n in range(anchor_len, 1, -1):
            anchor = chunk_tokens[:n]
            anchor_pos = self._find_token_sequence(raw_tokens, anchor, start_token_idx)
            if anchor_pos is not None:
                anchor_len = n
                break

        if anchor_pos is None:
            return -1, -1, start_token_idx

        tail_anchor = chunk_tokens[-anchor_len:]
        tail_pos = self._find_token_sequence(raw_tokens, tail_anchor, anchor_pos + anchor_len)

        if tail_pos is not None:
            end_token_idx = tail_pos + anchor_len - 1
        else:
            end_token_idx = min(len(raw_tokens) - 1, anchor_pos + len(chunk_tokens) - 1)

        start_char = raw_spans[anchor_pos][0]
        end_char = raw_spans[end_token_idx][1]
        return start_char, end_char, end_token_idx + 1

    def _detect_paywalled(self, text: str, fetch_meta: Optional[Dict[str, Any]]) -> bool:
        if fetch_meta:
            status = fetch_meta.get("status_code")
            if status in (402, 403, 401, 451):
                return True
        if not text:
            return False
        paywall_terms = ["subscribe", "subscription", "sign in to continue", "paywall", "purchase access"]
        text_lc = text.lower()
        return any(t in text_lc for t in paywall_terms)

    def _is_anti_bot_content(self, text: str) -> bool:
        if not text:
            return False
        t = text.lower()
        patterns = [
            "access to this page has been denied",
            "verify you are human",
            "captcha",
            "cloudflare",
            "attention required",
            "access denied",
            "proof of work",
            "anubis",
            "jshelter",
            "press & hold",
            "press and hold",
            "confirm you are a human",
            "checking your browser",
            "security check",
            "not a bot",
            "making sure you're not a bot",
        ]
        return any(p in t for p in patterns)

    @staticmethod
    def _extract_domain_from_url(url: str) -> str:
        raw = (url or "").strip()
        if not raw:
            return ""
        try:
            parsed = urlparse(raw if "://" in raw else f"https://{raw}")
            domain = (parsed.netloc or "").lower().strip()
            if domain.startswith("www."):
                domain = domain[4:]
            return domain
        except Exception:
            return ""

    def _intent_domain_quality(self, domain: str) -> int:
        d = (domain or "").lower().strip()
        if not d:
            return 0
        if any(d == bad or d.endswith(f".{bad}") for bad in self.intent_low_value_domains):
            return -2
        if d.endswith(".gov") or d.endswith(".edu"):
            return 2
        if any(h in d for h in self.intent_high_value_domain_hints):
            return 1
        return 0

    def _infer_source_intents(
        self,
        *,
        source_url: str,
        resolved_url: str,
        source_type: str,
        link_text: str,
        content_text: str,
    ) -> Dict[str, Any]:
        source_type_lc = (source_type or "").lower().strip()
        domain = self._extract_domain_from_url(resolved_url or source_url)
        domain_quality = self._intent_domain_quality(domain)

        url_blob = " ".join(
            [
                (source_url or ""),
                (resolved_url or ""),
                (link_text or ""),
                source_type_lc,
            ]
        ).lower()
        text_blob = (content_text or "")[: self.intent_scan_text_chars].lower()

        scores: Dict[str, float] = {}
        hits: Dict[str, int] = {}
        for intent, patterns in self.intent_patterns.items():
            url_hits = 0
            text_hits = 0
            for pat in patterns:
                if pat.search(url_blob):
                    url_hits += 1
                if pat.search(text_blob):
                    text_hits += 1
            type_hit = int(any(h in source_type_lc for h in self.intent_source_type_hints.get(intent, set())))
            score = (
                0.30 * float(url_hits)
                + 0.25 * float(text_hits)
                + 0.20 * float(type_hit)
                + 0.20 * float(max(0, domain_quality))
            )
            if domain_quality < 0:
                score -= 0.25
            score = min(1.0, max(0.0, score))
            if score > 0:
                scores[intent] = round(score, 4)
                hits[intent] = int(url_hits + text_hits + type_hit)

        sorted_intents = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        top_intents = [k for k, _ in sorted_intents]
        required_hits = [k for k in top_intents if k in self.intent_required_set]
        max_required_conf = max([scores[k] for k in required_hits], default=0.0)

        gate_pass = bool(required_hits and max_required_conf >= self.intent_min_source_confidence)
        gate_reason = ""
        if not required_hits:
            gate_reason = "no_required_intent_match"
        elif max_required_conf < self.intent_min_source_confidence:
            gate_reason = f"low_intent_confidence:{max_required_conf:.2f}"
        if gate_pass and self.intent_strict_domain_filter and domain_quality < 0:
            gate_pass = False
            gate_reason = "low_value_domain_strict_filter"

        return {
            "domain": domain,
            "domain_quality": domain_quality,
            "intent_scores": scores,
            "intent_hits": hits,
            "intents": top_intents,
            "required_intents": required_hits,
            "primary_intent": top_intents[0] if top_intents else "",
            "max_required_confidence": round(max_required_conf, 4),
            "gate_pass": gate_pass,
            "gate_reason": gate_reason,
        }

    def _classify_chunk_intent(self, text: str) -> List[str]:
        """Return at most two intent tags for a chunk, ordered by hit count.

        Replaces the previous "tag every chunk with the union of required
        intents" behaviour. We re-use the same regex patterns used at the
        source level but score per-chunk so that, e.g., a Du Bois bio
        paragraph mentioning a 1909 speech gets only ``biography`` (and
        possibly ``speech``) rather than all five required types.
        """
        if not text:
            return []
        sample = text[: self.intent_scan_text_chars].lower()
        scores: List[Tuple[str, int]] = []
        for intent, patterns in self.intent_patterns.items():
            hits = 0
            for pat in patterns:
                hits += len(pat.findall(sample))
            if hits > 0:
                scores.append((intent, hits))
        if not scores:
            return []
        scores.sort(key=lambda kv: kv[1], reverse=True)
        return [k for k, _ in scores[:2]]

    @staticmethod
    def _mark_duplicate_chunks(source_chunks: List[Dict[str, Any]]) -> None:
        chunk_hash_map: Dict[str, str] = {}
        for ch in source_chunks:
            t_hash = ch.get("text_hash")
            if not t_hash:
                continue
            if t_hash in chunk_hash_map:
                ch["duplicate_of_chunk_id"] = chunk_hash_map[t_hash]
            else:
                chunk_hash_map[t_hash] = ch.get("chunk_id", "")

    def _build_intent_summary(
        self,
        *,
        sources_meta: List[Dict[str, Any]],
        source_chunks: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        source_counts = Counter()
        chunk_counts = Counter()
        required_source_counts = Counter()
        required_chunk_counts = Counter()

        accepted_sources: List[Dict[str, Any]] = []
        for s in sources_meta:
            if s.get("status") == "blocked":
                continue
            if self.intent_gating_enabled and not bool(s.get("intent_gate_pass")):
                continue
            accepted_sources.append(s)

        for s in accepted_sources:
            intents = s.get("intent_tags") or []
            for intent in intents:
                source_counts[str(intent)] += 1
            req = s.get("intent_required_hits") or []
            for intent in req:
                required_source_counts[str(intent)] += 1

        for ch in source_chunks:
            intents = ch.get("intent_tags") or []
            for intent in intents:
                chunk_counts[str(intent)] += 1
                if intent in self.intent_required_set:
                    required_chunk_counts[str(intent)] += 1

        covered_required = [
            intent
            for intent in self.intent_required_types
            if required_source_counts.get(intent, 0) > 0
        ]

        return {
            "enabled": bool(self.intent_gating_enabled),
            "required_types": list(self.intent_required_types),
            "required_types_covered": sorted(covered_required),
            "required_types_covered_count": len(covered_required),
            "accepted_source_count": len(accepted_sources),
            "accepted_chunk_count": len(source_chunks),
            "source_counts": dict(source_counts),
            "chunk_counts": dict(chunk_counts),
            "required_source_counts": dict(required_source_counts),
            "required_chunk_counts": dict(required_chunk_counts),
            "min_covered_types": self.intent_min_covered_types,
            "min_sources": self.intent_min_sources,
            "min_chunks": self.intent_min_chunks,
            "min_source_confidence": self.intent_min_source_confidence,
        }

    def _passes_profile_intent_gate(self, intent_summary: Dict[str, Any]) -> Tuple[bool, str]:
        if not self.intent_gating_enabled:
            return True, ""
        covered = int(intent_summary.get("required_types_covered_count", 0))
        sources = int(intent_summary.get("accepted_source_count", 0))
        chunks = int(intent_summary.get("accepted_chunk_count", 0))

        if covered < self.intent_min_covered_types:
            return (
                False,
                f"intent_gate_failed:covered_types_{covered}_lt_{self.intent_min_covered_types}",
            )
        if sources < self.intent_min_sources:
            return (
                False,
                f"intent_gate_failed:accepted_sources_{sources}_lt_{self.intent_min_sources}",
            )
        if chunks < self.intent_min_chunks:
            return (
                False,
                f"intent_gate_failed:accepted_chunks_{chunks}_lt_{self.intent_min_chunks}",
            )
        return True, ""

    def _process_sources_with_intent_gates(
        self,
        *,
        source_items: List[Dict[str, Any]],
        profile_name: str,
        profile_id: str,
        profile_url: str,
        search_meta: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        source_chunks: List[Dict[str, Any]] = []
        sources_meta: List[Dict[str, Any]] = []
        allowed_source_items: List[Dict[str, Any]] = []
        seen_source_ids: set = set()

        for item in source_items:
            source_url = item.get("source_url", "") or ""
            resolved_url = item.get("resolved_url", "") or source_url
            source_type = item.get("source_type", "webpage")
            content_text = item.get("content_text", "")
            fetch_meta = item.get("fetch_metadata", {})
            link_text = item.get("link_text") or extract_title_from_text(content_text)

            # Skip known UI-chrome / paste-site / social hosts before they ever
            # get registered. Previously help.pbs.org/support pages, scribd
            # listing pages and pdfcoffee thumbnails became "biography"
            # sources by accident.
            if is_noise_domain(resolved_url) or is_noise_domain(source_url):
                continue

            # Drop content that does not mention the subject at all when we
            # have a confident name — protects against co-honoree contamination
            # (the Owens "Byrd's Profile" / "Marshall's Profile" leak) and
            # against ingesting an author's own book as if it were biography.
            if profile_name and not self._source_matches_target_profile(
                source_type=source_type,
                source_url=resolved_url or source_url,
                link_text=link_text,
                content_text=content_text,
                profile_name=profile_name,
            ):
                continue

            record = self._register_source(
                source_url=source_url,
                resolved_url=resolved_url,
                source_type=source_type,
                content_text=content_text,
                fetch_meta=fetch_meta,
                profile_id=profile_id,
                profile_url=profile_url,
                search_meta=search_meta,
            )

            intent_meta = self._infer_source_intents(
                source_url=source_url,
                resolved_url=resolved_url,
                source_type=source_type,
                link_text=link_text,
                content_text=content_text,
            )
            raw_intent_gate_pass = bool(intent_meta.get("gate_pass"))
            effective_intent_gate_pass = raw_intent_gate_pass if self.intent_gating_enabled else True

            already_seen = record.source_id in seen_source_ids
            if not already_seen:
                # ``link_text`` (page title) is what the chunker uses to
                # pick a human-readable section label. Falling back to the
                # extracted-from-text title means downstream consumers
                # cite something like "W. E. B. Du Bois biography" rather
                # than "34807" or a DOI hash.
                page_title = self._derive_page_title(link_text, content_text)
                sources_meta.append(
                    {
                        "source_id": record.source_id,
                        "source_url": source_url,
                        "resolved_url": record.final_url,
                        "source_type": source_type,
                        "link_text": page_title,
                        "status": record.status,
                        "allowed_use": record.allowed_use,
                        "license_type": record.license_type,
                        "robots_allowed": record.robots_allowed,
                        "paywalled": record.paywalled,
                        "intent_tags": intent_meta.get("intents") or [],
                        "intent_required_hits": intent_meta.get("required_intents") or [],
                        "intent_primary": intent_meta.get("primary_intent", ""),
                        "intent_scores": intent_meta.get("intent_scores") or {},
                        "intent_confidence": intent_meta.get("max_required_confidence", 0.0),
                        "intent_domain": intent_meta.get("domain", ""),
                        "intent_domain_quality": intent_meta.get("domain_quality", 0),
                        "intent_raw_gate_pass": raw_intent_gate_pass,
                        "intent_gate_pass": effective_intent_gate_pass,
                        "intent_gate_enforced": bool(self.intent_gating_enabled),
                        "intent_gate_reason": intent_meta.get("gate_reason", ""),
                    }
                )
                seen_source_ids.add(record.source_id)
            else:
                continue

            if record.status == "blocked":
                continue
            if not effective_intent_gate_pass:
                continue

            allowed_source_items.append(item)
            lang = "unknown"
            if record.content and isinstance(record.content, dict):
                lang = record.content.get("language", "unknown") or "unknown"
            _, chunks = self._chunk_source_text(
                source_id=record.source_id,
                source_url=source_url,
                source_type=source_type,
                content_text=content_text,
                profile_name=profile_name,
                allowed_use=record.allowed_use,
                language=lang,
                domain_quality=int(intent_meta.get("domain_quality", 0) or 0),
            )
            primary_intent = intent_meta.get("primary_intent", "") or ""
            # Tag chunks with the *single* primary intent inherited from the
            # source plus whatever per-chunk pattern hits exist. This
            # replaces the previous behaviour of stamping every chunk with
            # the full required-types set, which made intent-based retrieval
            # filtering meaningless (every chunk was speech+interview+paper+
            # testimony+policy_statement).
            for ch in chunks:
                chunk_intents = self._classify_chunk_intent(ch.get("text", ""))
                if primary_intent and primary_intent not in chunk_intents:
                    chunk_intents.insert(0, primary_intent)
                ch["intent_tags"] = chunk_intents
                ch["intent_primary"] = chunk_intents[0] if chunk_intents else primary_intent
                ch["intent_confidence"] = intent_meta.get("max_required_confidence", 0.0)
                ch["intent_gate_pass"] = effective_intent_gate_pass
            source_chunks.extend(chunks)

        self._mark_duplicate_chunks(source_chunks)
        intent_summary = self._build_intent_summary(
            sources_meta=sources_meta,
            source_chunks=source_chunks,
        )
        gate_pass, gate_reason = self._passes_profile_intent_gate(intent_summary)
        intent_summary["profile_gate_pass"] = bool(gate_pass)
        intent_summary["profile_gate_reason"] = gate_reason

        return {
            "source_chunks": source_chunks,
            "sources_meta": sources_meta,
            "allowed_source_items": allowed_source_items,
            "intent_summary": intent_summary,
            "profile_gate_pass": gate_pass,
            "profile_gate_reason": gate_reason,
        }

    def _decide_source_status(
        self,
        license_type: str,
        allowed_use: str,
        robots_allowed: Optional[bool],
        paywalled: bool
    ) -> Tuple[str, str]:
        """Promote sources to ``allowed`` when license/usage is known.

        Previous behaviour treated every source as ``review`` because all
        sources arrived with ``license_type=unknown`` and the original gate
        also (incorrectly) flagged ``allowed_use != "facts_only"`` as a
        rejection — which inverted the ladder ``facts_only`` < ``short_quotes``
        < ``full_text``. With ``infer_license_for_url`` populating real values,
        the rule is now: sources with a known permissive license, robots OK,
        and no paywall are promoted to ``allowed`` and become quote-eligible.
        """
        reasons = []
        if robots_allowed is False:
            reasons.append("robots_disallow")
        if paywalled:
            reasons.append("paywalled")
        if license_type == "unknown":
            reasons.append("license_unknown")
        if allowed_use not in ("facts_only", "short_quotes", "full_text"):
            reasons.append("allowed_use_unrecognized")
        if reasons:
            status = "blocked" if self.strict_source_policy else "review"
            return status, ",".join(reasons)
        return "allowed", ""

    def _register_source(
        self,
        source_url: str,
        resolved_url: str,
        source_type: str,
        content_text: str,
        fetch_meta: Optional[Dict[str, Any]],
        profile_id: str,
        profile_url: str,
        search_meta: Optional[Dict[str, Any]] = None
    ) -> SourceRecord:
        canonical = normalize_url(resolved_url or source_url)
        if not canonical:
            canonical = f"content:{compute_text_hash(content_text)[:16]}"
        source_id = make_source_id(canonical)
        domain = ""
        try:
            domain = urlparse(canonical).netloc
        except Exception:
            domain = ""

        # License/usage inference: previously every source was hard-coded to
        # ``unknown`` / ``facts_only``, which forced ``quote_ok=False`` for the
        # entire corpus (including loc.gov / nps.gov / archive.org content).
        license_info = infer_license_for_url(canonical or source_url)
        license_type = license_info.get("license_type", "unknown")
        allowed_use = license_info.get("allowed_use", "facts_only")
        license_url = license_info.get("license_url", "")
        rights_holder = license_info.get("rights_holder", "")
        robots_allowed = self.robots_cache.allowed(canonical) if canonical else None
        paywalled = self._detect_paywalled(content_text, fetch_meta)
        status, status_reason = self._decide_source_status(license_type, allowed_use, robots_allowed, paywalled)
        if self._is_anti_bot_content(content_text):
            status = "blocked"
            status_reason = ",".join([r for r in [status_reason, "anti_bot_content"] if r])

        lang, lang_conf = detect_language(content_text)
        quality = compute_quality_metrics(content_text)
        pii = detect_pii(content_text)
        text_hash = compute_text_hash(content_text)
        # Basic simhash signature (tokens)
        tokens = re.findall(r"[A-Za-z0-9]+", content_text.lower())
        simhash = compute_simhash(tokens) if tokens else None

        duplicate_of = self._content_hash_index.get(text_hash)
        if not duplicate_of:
            self._content_hash_index[text_hash] = source_id

        if not search_meta:
            search_meta = self.default_search_meta

        record = SourceRecord(
            source_id=source_id,
            url=source_url,
            accessed_at=datetime.now(timezone.utc).isoformat(),
            domain=domain,
            license_type=license_type,
            license_url=license_url,
            rights_holder=rights_holder,
            allowed_use=allowed_use,
            paywalled=paywalled,
            robots_allowed=robots_allowed,
            copyright_notes="",
            status=status,
            status_reason=status_reason,
            final_url=resolved_url or source_url,
            fetch=fetch_meta or {},
            content={
                "text_hash": text_hash,
                "text_length": len(content_text or ""),
                "language": lang,
                "language_confidence": round(lang_conf, 4),
                "quality": quality,
                "pii": pii,
                "simhash": simhash,
                "duplicate_of": duplicate_of,
            },
            search=search_meta or {},
            profile_context={
                "profile_id": profile_id,
                "profile_url": profile_url,
                "source_type": source_type,
            },
        )
        self.source_registry.upsert(record)
        return record

    def _chunk_source_text(
        self,
        source_id: str,
        source_url: str,
        source_type: str,
        content_text: str,
        profile_name: str,
        allowed_use: str,
        language: str,
        domain_quality: int = 0,
    ) -> Tuple[str, List[Dict[str, Any]]]:
        """
        Clean and chunk a single source, returning (cleaned_text, chunk_records).

        Per-source LLM cleaning is intentionally skipped: it rewrites the
        text and breaks raw_text → chunk offset mapping for ~70% of chunks.
        The downstream chunk-level filters (boilerplate, subject mention,
        domain-quality) handle noise without needing the LLM pass here.
        """
        if not content_text or not content_text.strip():
            return "", []

        # Light per-line filter is still applied, but only when the project
        # explicitly asked for it (PROFILE_TEXT_FILTER_ENABLED=1).
        filtered_text = self._filter_text_for_profile(content_text, profile_name)

        chunks = self.cleaning_service.clean_and_chunk_text(
            text=filtered_text,
            profile_url=source_url,
            section_header=""
        )
        cleaned_text = " ".join([c.get("text", "") for c in chunks if c.get("text")]).strip()

        raw_index = self._build_raw_token_index(content_text)
        raw_search_start = 0
        chunk_records: List[Dict[str, Any]] = []
        # Track recent in-source chunk window so we can keep pronoun-only
        # chunks that immediately follow a subject-mentioning chunk (the
        # subject is still in scope). The page-level subject filter has
        # already cleared the page; this loop drops the leftover *chunks*
        # inside a multi-subject page (the Owens co-honoree booklet ships
        # ~150 paragraphs; only ~25 are about Owens).
        last_subject_chunk_idx = -10
        for idx, ch in enumerate(chunks):
            chunk_text = (ch.get("text") or "").strip()
            if not chunk_text:
                continue
            if self._chunk_is_boilerplate(chunk_text):
                continue
            mentions_subject = self._chunk_mentions_subject(chunk_text, profile_name)
            if mentions_subject:
                last_subject_chunk_idx = idx
            else:
                # No mention of subject in this chunk.
                # Low-quality domains: drop unconditionally.
                if domain_quality < 0:
                    continue
                # Other domains: allow only if immediately adjacent to a
                # subject-mentioning chunk AND looks like a pronoun
                # continuation. Otherwise drop — this is co-honoree /
                # anthology / catalog contamination.
                if (idx - last_subject_chunk_idx) > 1:
                    continue
                if not self._chunk_is_pronoun_continuation(chunk_text, profile_name):
                    continue
            raw_start, raw_end, raw_search_start = self._map_chunk_offsets(
                raw_index, chunk_text, raw_search_start
            )
            # Lenient fallback: try a shorter prefix match against the raw
            # text so chunks that were mildly reformatted by the cleaner
            # still get an offset_start instead of -1.
            if raw_start < 0 and raw_index and len(chunk_text) >= 80:
                try:
                    needle = re.sub(r"\s+", " ", chunk_text[:80]).strip().lower()
                    haystack = re.sub(r"\s+", " ", raw_index["text"]).lower()
                    pos = haystack.find(needle)
                    if pos >= 0:
                        raw_start = pos
                        raw_end = pos + len(needle)
                except Exception:
                    pass
            chunk_id = f"{source_id}:{idx}"
            text_hash = compute_text_hash(chunk_text)
            chunk_records.append({
                "chunk_id": chunk_id,
                "source_id": source_id,
                "source_url": source_url,
                "source_type": source_type,
                "text": chunk_text,
                "text_hash": text_hash,
                "offset_start": raw_start,
                "offset_end": raw_end,
                "allowed_use": allowed_use,
                "quote_ok": allowed_use in ("short_quotes", "full_text"),
                "is_summary": False,
                "language": language,
                "subject_mention": mentions_subject,
                "domain_quality": int(domain_quality),
            })

        return cleaned_text, chunk_records

    @staticmethod
    def _derive_page_title(link_text: str, content_text: str) -> str:
        """Pick the best human-readable title for a source.

        Prefers the scraper's link_text/title, then the first non-noise
        line of content. Strips chrome tokens, trailing site names after
        " | " or " - " separators, and obvious ID/hash fragments.
        """
        candidates: List[str] = []
        if link_text:
            candidates.append(link_text)
        # First few non-empty lines of cleaned content as fallback titles.
        if content_text:
            count = 0
            for ln in content_text.splitlines():
                ln = ln.strip()
                if ln and len(ln) >= 6 and not re.match(r"^[\W_]+$", ln):
                    candidates.append(ln)
                    count += 1
                    if count >= 3:
                        break
        for raw in candidates:
            title = raw.strip()
            # Chop trailing " | Site Name" / " - Site Name" suffixes.
            for sep in (" | ", " — ", " – ", " - "):
                if sep in title:
                    head = title.split(sep, 1)[0].strip()
                    if head and len(head) >= 5:
                        title = head
                        break
            # Drop colon-separated breadcrumb tails like
            # "Kevin Brown: Directory: Faculty: About us: Maurer School of Law".
            if title.count(":") >= 2:
                segs = [s.strip() for s in title.split(":") if s.strip()]
                if len(segs) >= 3:
                    head = segs[0]
                    tail_avg = sum(len(s) for s in segs[1:]) / max(1, len(segs) - 1)
                    if head and len(head) >= 4 and tail_avg <= 40:
                        title = head
            title = re.sub(r"\s+", " ", title).strip()
            if not title:
                continue
            if len(title) < 5 or len(title) > 200:
                continue
            lower = title.lower()
            if lower in {
                "loading", "shopping basket", "shopping cart", "search",
                "menu", "home", "sign in", "log in", "subscribe",
            }:
                continue
            if re.match(r"^[A-Fa-f0-9\-]{16,}$", title.replace(" ", "")):
                continue
            if re.match(r"^[0-9]{1,8}$", title):
                continue
            return title[:120]
        return ""

    @staticmethod
    def _chunk_is_boilerplate(text: str) -> bool:
        """Drop chunks that are video lists, table-of-contents, or
        otherwise low-information page chrome — these were clogging
        the Woodson `videos` page and many archival index pages."""
        if not text:
            return True
        stripped = text.strip()
        if len(stripped) < 120:
            return True
        lines = [ln.strip() for ln in stripped.splitlines() if ln.strip()]
        if not lines:
            return True
        # Mostly bullet/listy lines (common in nav menus, video catalogues).
        listy = sum(
            1 for ln in lines
            if ln.startswith(("•", "-", "*", "·"))
            or len(ln.split()) <= 4
            or re.match(r"^\(\d{4}\)$", ln)
        )
        if listy / max(1, len(lines)) >= 0.6 and len(lines) >= 6:
            return True
        # Mostly non-letters (TOC, dotted leaders, page numbers).
        letters = sum(1 for ch in stripped if ch.isalpha())
        if letters / max(1, len(stripped)) < 0.5:
            return True
        return False

    @staticmethod
    def _chunk_mentions_subject(text: str, profile_name: str) -> bool:
        """True if a chunk explicitly names the subject AND no other person
        clearly dominates the paragraph.

        Originally this was a permissive single-hit check, which let
        co-faculty / co-author paragraphs through whenever the IU
        directory page or academia.edu page also happened to name the
        subject once. We now also reject chunks where the *dominant*
        narrative subject — measured by how many sentences a person's
        full Title-Case name leads with a verb pattern — is someone
        else. This kills the Duncan↔Doerrenberg and Nesbitt↔Amaral
        contamination clusters.
        """
        if not text or not profile_name:
            return False
        text_lc = text.lower()
        name = profile_name.strip().lower()
        parts = [p for p in name.split() if len(p) >= 3]
        if not parts:
            return False
        last = parts[-1]
        first = parts[0]

        subject_hit = False
        if name and name in text_lc:
            subject_hit = True
        elif last and last in text_lc:
            subject_hit = True
        elif first and first in text_lc and any(p in text_lc for p in parts[1:]):
            subject_hit = True

        if not subject_hit:
            return False

        # Count "X is/was/wrote/serves/teaches/argues/studies" leading
        # patterns for any Title-Case full name. The subject's own name
        # parts go in a stop-set so they don't count against themselves.
        narrative_pat = re.compile(
            r"(?:^|[\.\n!?])\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})"
            r"\s+(?:is|was|wrote|serves?|served|teaches?|taught|argues?|argued|"
            r"studies|studied|founded|leads?|led|directs?|directed|holds?|"
            r"earned|graduated|published|received|focuses?|works?|joined)\b"
        )
        leaders = [m.group(1) for m in narrative_pat.finditer(text)]
        if not leaders:
            return True

        from collections import Counter
        counter: Counter = Counter()
        for ld in leaders:
            ld_lc = ld.lower()
            # Treat any name fragment that overlaps with the subject's
            # parts as the subject (e.g. "Du Bois" / "W. E. B. Du Bois").
            if last and last in ld_lc:
                counter["__subject__"] += 1
            elif name and name in ld_lc:
                counter["__subject__"] += 1
            else:
                counter[ld_lc] += 1

        subj_count = counter.pop("__subject__", 0)
        if not counter:
            return True
        top_other_count = max(counter.values())
        # Reject if some other person leads ≥2 sentences AND outpaces
        # the subject by 2+. A single passing reference is fine; a
        # paragraph clearly *about* somebody else is not.
        if top_other_count >= 2 and (top_other_count - subj_count) >= 2:
            return False
        return True

    @staticmethod
    def _chunk_is_pronoun_continuation(text: str, profile_name: str) -> bool:
        """Heuristic: chunk has no subject mention but does contain
        third-person pronouns and no other named person leading a sentence —
        so it's likely a continuation paragraph about the subject."""
        if not text:
            return False
        # If another full proper-noun name leads a sentence ("Smith was..."),
        # treat as a different subject's paragraph.
        other_subject = re.search(
            r"(?:^|[\.\n])\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s+(?:was|is|earned|served|founded|published|received|wrote|graduated|holds|completed)\b",
            text,
        )
        if other_subject:
            cand = other_subject.group(1).lower()
            name_lc = (profile_name or "").lower()
            if cand and cand not in name_lc and not any(p in name_lc for p in cand.split()):
                return False
        text_lc = text.lower()
        return bool(re.search(r"\b(he|she|his|her|him|hers)\b", text_lc))

    def _save_source_chunks(
        self,
        profile_id: str,
        source_chunks: List[Dict[str, Any]],
        sources_meta: List[Dict[str, Any]],
        intent_summary: Optional[Dict[str, Any]] = None
    ) -> Path:
        profile_dir = self.output_dir / "profiles" / profile_id
        profile_dir.mkdir(parents=True, exist_ok=True)
        output_path = profile_dir / "source_chunks.json"
        payload = {
            "profile_id": profile_id,
            "sources": sources_meta,
            "chunks": source_chunks,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        if intent_summary is not None:
            payload["intent_gates"] = intent_summary
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return output_path

    def _save_claims(
        self,
        profile_id: str,
        claims: List[Dict[str, Any]]
    ) -> Path:
        profile_dir = self.output_dir / "profiles" / profile_id
        profile_dir.mkdir(parents=True, exist_ok=True)
        output_path = profile_dir / "claims.json"
        payload = {
            "profile_id": profile_id,
            "claims": claims,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        return output_path
    
    async def initialize_scraper(self):
        """Initialize scraper (async)"""
        if self.scraper is None:
            self.scraper = await get_scraper()
            print("[Pipeline] Scraper initialized")
    
    def read_excel(self, excel_path: str) -> pd.DataFrame:
        """
        Read Excel file and extract profile URLs
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            DataFrame with profile URLs
        """
        print(f"\n[Step 1] Reading Excel file: {excel_path}")
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        df = pd.read_excel(excel_path)
        
        # Validate and find URL column
        if 'source' in df.columns and 'profile_url' not in df.columns:
            df.rename(columns={'source': 'profile_url'}, inplace=True)
        elif 'profile_url' not in df.columns:
            url_columns = [col for col in df.columns if 'url' in col.lower() or 'link' in col.lower()]
            if url_columns:
                df.rename(columns={url_columns[0]: 'profile_url'}, inplace=True)
            else:
                raise ValueError("Excel file must contain a 'source' or 'profile_url' column")
        
        # Normalize 'name' column (case-insensitive)
        name_columns = [col for col in df.columns if col.lower() == 'name']
        if name_columns and 'name' not in df.columns:
            # Rename first name column to 'name' for consistency
            df.rename(columns={name_columns[0]: 'name'}, inplace=True)
        
        # Filter valid URLs and preserve original Excel row number (header row is 1)
        valid_df = df[df['profile_url'].notna()].copy()
        valid_df["_excel_row_number"] = valid_df.index + 2
        valid_df = valid_df.reset_index(drop=True)
        
        if valid_df.empty:
            raise ValueError("No valid URLs found in Excel file")
        
        # Log name column status
        if 'name' in valid_df.columns:
            names_found = valid_df['name'].notna().sum()
            print(f"[Step 1] Found {len(valid_df)} valid profile URLs")
            print(f"[Step 1] Found 'name' column with {names_found} names")
        else:
            print(f"[Step 1] Found {len(valid_df)} valid profile URLs")
            print(f"[Step 1] Warning: no 'name' column found - will use scraped names")
        
        return valid_df

    def _open_excel_ignore_tracker(self, excel_path: str) -> Optional[Dict[str, Any]]:
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook[workbook.sheetnames[0]]

            headers: Dict[str, int] = {}
            for col_idx in range(1, sheet.max_column + 1):
                value = sheet.cell(row=1, column=col_idx).value
                if value is None:
                    continue
                headers[str(value).strip().lower()] = col_idx

            def ensure_column(header_name: str) -> int:
                key = header_name.strip().lower()
                existing = headers.get(key)
                if existing is not None:
                    return int(existing)
                new_col = sheet.max_column + 1
                sheet.cell(row=1, column=new_col, value=header_name)
                headers[key] = new_col
                return int(new_col)

            ignored_col = ensure_column("ignored")
            ignored_reason_col = ensure_column("ignored_reason")
            scrape_status_col = ensure_column("scrape_status")
            scrape_issue_col = ensure_column("scrape_issue")

            return {
                "workbook": workbook,
                "sheet": sheet,
                "ignored_col": ignored_col,
                "ignored_reason_col": ignored_reason_col,
                "scrape_status_col": scrape_status_col,
                "scrape_issue_col": scrape_issue_col,
            }
        except Exception as e:
            print(f"[Excel] Warning: unable to prepare scrape-status columns: {e}")
            self._log_error(f"[excel_ignore_tracker] init failed :: {e}")
            return None

    def _set_excel_ignore_status(
        self,
        tracker: Optional[Dict[str, Any]],
        *,
        excel_row_number: Optional[int],
        ignored: bool,
        reason: str = "",
        status: str = "",
        scrape_issue: str = "",
    ) -> None:
        if not tracker:
            return
        if excel_row_number is None or excel_row_number < 2:
            return

        sheet = tracker["sheet"]
        sheet.cell(
            row=excel_row_number,
            column=int(tracker["ignored_col"]),
            value="ignored" if ignored else "",
        )
        sheet.cell(
            row=excel_row_number,
            column=int(tracker["ignored_reason_col"]),
            value=(reason or "") if ignored else "",
        )
        status_value = self._normalize_inline_whitespace(status).lower()
        if status_value not in {"success", "failed", "ignored"}:
            status_value = self._normalize_inline_whitespace(status)
        sheet.cell(
            row=excel_row_number,
            column=int(tracker["scrape_status_col"]),
            value=status_value,
        )
        issue_value = self._normalize_inline_whitespace(scrape_issue) if status_value == "failed" else ""
        sheet.cell(
            row=excel_row_number,
            column=int(tracker["scrape_issue_col"]),
            value=issue_value,
        )

    def _save_excel_ignore_tracker(self, tracker: Optional[Dict[str, Any]], excel_path: str) -> None:
        if not tracker:
            return
        workbook = tracker.get("workbook")
        if workbook is None:
            return
        try:
            workbook.save(excel_path)
            print(f"[Excel] Updated scrape status in: {excel_path}")
        except Exception as e:
            print(f"[Excel] Warning: failed to save scrape status updates: {e}")
            self._log_error(f"[excel_ignore_tracker] save failed :: {e}")
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _set_incremental_sync_error(self, message: str) -> None:
        self._incremental_sync_stats["last_error"] = message
        self._log_error(f"[incremental_sync] {message}")

    def _init_incremental_vector_services(self) -> bool:
        if self.incremental_skip_pinecone:
            return False
        if self._incremental_vector_db is not None and self._incremental_embeddings_service is not None:
            return True
        try:
            from api.services.vector_db import get_vector_db
            from api.services.embeddings import get_embeddings_service
            from config.pinecone_config import INDEX_NAME, INDEX_DIMENSION

            self._incremental_vector_db = get_vector_db(index_name=INDEX_NAME, dimension=INDEX_DIMENSION)
            self._incremental_embeddings_service = get_embeddings_service()
            self._incremental_index_dimension = int(INDEX_DIMENSION)
            return True
        except Exception as e:
            self._set_incremental_sync_error(f"vector_service_init_failed :: {e}")
            self._incremental_vector_db = None
            self._incremental_embeddings_service = None
            return False

    def _init_incremental_mongo_sync(self) -> bool:
        if self.incremental_skip_mongo:
            return False
        if self._incremental_mongo_sync is not None:
            return True
        try:
            from sync_profiles_to_mongodb import MongoDBScholarSync

            self._incremental_mongo_sync = MongoDBScholarSync()
            if not self.incremental_skip_indexes and not self._incremental_indexes_created:
                self._incremental_mongo_sync.create_indexes()
                self._incremental_indexes_created = True
            return True
        except Exception as e:
            self._set_incremental_sync_error(f"mongo_sync_init_failed :: {e}")
            self._incremental_mongo_sync = None
            return False

    def _load_sync_chunks_for_profile(
        self,
        *,
        profile_id: str,
        profile_name: str,
        chunks_path: Optional[str],
    ) -> List[Dict[str, Any]]:
        if not chunks_path:
            return []
        path = Path(chunks_path)
        if not path.exists():
            self._set_incremental_sync_error(f"chunks_path_missing :: {path}")
            return []
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as e:
            self._set_incremental_sync_error(f"chunks_path_read_failed :: {path} :: {e}")
            return []

        sections = payload.get("sections") or {}
        if not isinstance(sections, dict):
            return []

        pid = str(payload.get("profile_id") or profile_id or "").strip()
        if not pid:
            return []
        pname = (profile_name or "Unknown").strip() or "Unknown"
        rows: List[Dict[str, Any]] = []
        for section_name, section_chunks in sections.items():
            if not isinstance(section_chunks, list):
                continue
            for idx, chunk in enumerate(section_chunks):
                if not isinstance(chunk, dict):
                    continue
                text = chunk.get("text", "")
                if not isinstance(text, str) or not text.strip():
                    continue
                order = chunk.get("order", chunk.get("chunk_index", idx))
                chunk_id = str(chunk.get("chunk_id") or "").strip()
                if not chunk_id:
                    chunk_id = compute_text_hash(f"{pid}|{section_name}|{order}|{text[:250]}")[:24]
                rows.append(
                    {
                        "profile_id": pid,
                        "professor_id": pid,
                        "professor_name": pname,
                        "section": str(chunk.get("section") or section_name or "Unknown"),
                        "chunk_id": chunk_id,
                        "order": int(order) if isinstance(order, (int, float)) else idx,
                        "text": text,
                    }
                )
        return rows

    def _upload_chunks_batch_incremental(self, chunks: List[Dict[str, Any]]) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "uploaded_vectors": 0,
            "failed_vectors": 0,
            "uploaded_profile_ids": set(),
        }
        if self.incremental_skip_pinecone:
            return summary
        if not chunks:
            return summary
        if not self._init_incremental_vector_services():
            summary["failed_vectors"] = len(chunks)
            return summary

        dim = int(self._incremental_index_dimension or 0)
        if dim <= 0:
            self._set_incremental_sync_error("pinecone_dimension_invalid")
            summary["failed_vectors"] = len(chunks)
            return summary

        valid_chunks = [c for c in chunks if (c.get("text") or "").strip()]
        summary["failed_vectors"] += max(0, len(chunks) - len(valid_chunks))

        for batch_start in range(0, len(valid_chunks), self.incremental_pinecone_batch_size):
            batch = valid_chunks[batch_start : batch_start + self.incremental_pinecone_batch_size]
            texts = [str(c.get("text", "")) for c in batch]
            try:
                embeddings = self._incremental_embeddings_service.embed_batch(texts, batch_size=len(texts))
            except Exception as e:
                self._set_incremental_sync_error(f"pinecone_embedding_batch_failed :: {e}")
                summary["failed_vectors"] += len(batch)
                continue

            vectors: List[Dict[str, Any]] = []
            for i, chunk in enumerate(batch):
                if i >= len(embeddings):
                    summary["failed_vectors"] += 1
                    continue
                emb = embeddings[i]
                if not isinstance(emb, list):
                    summary["failed_vectors"] += 1
                    continue
                if len(emb) != dim or all(v == 0.0 for v in emb):
                    summary["failed_vectors"] += 1
                    continue

                profile_id = str(chunk.get("profile_id") or "")
                chunk_id = str(chunk.get("chunk_id") or "")
                if not profile_id:
                    summary["failed_vectors"] += 1
                    continue

                vector_id = f"chunk_{chunk_id}" if chunk_id else f"profile_{profile_id}_{i}"
                metadata = {
                    "profile_id": profile_id,
                    "professor_id": str(chunk.get("professor_id") or profile_id),
                    "professor_name": str(chunk.get("professor_name") or "Unknown"),
                    "section": str(chunk.get("section") or "Unknown"),
                    "chunk_id": chunk_id,
                    "order": (
                        int(chunk.get("order", 0))
                        if isinstance(chunk.get("order", 0), (int, float))
                        else 0
                    ),
                    "text": str(chunk.get("text") or ""),
                    "content_type": "profile_chunk",
                }
                vectors.append({"id": vector_id, "values": emb, "metadata": metadata})

            if not vectors:
                continue

            try:
                for j in range(0, len(vectors), 100):
                    sub = vectors[j : j + 100]
                    self._incremental_vector_db.index.upsert(vectors=sub)
                summary["uploaded_vectors"] += len(vectors)
                for vec in vectors:
                    pid = str(vec.get("metadata", {}).get("profile_id") or "")
                    if pid:
                        summary["uploaded_profile_ids"].add(pid)
            except Exception as e:
                self._set_incremental_sync_error(f"pinecone_upsert_failed :: {e}")
                summary["failed_vectors"] += len(vectors)

        return summary

    def _sync_profiles_batch_incremental(self, profile_name_map: Dict[str, str]) -> Dict[str, int]:
        summary = {"synced": 0, "failed": 0}
        if self.incremental_skip_mongo:
            return summary
        if not profile_name_map:
            return summary
        if not self._init_incremental_mongo_sync():
            summary["failed"] = len(profile_name_map)
            return summary

        for profile_id, profile_name in profile_name_map.items():
            try:
                ok = self._incremental_mongo_sync.sync_profile(profile_id, profile_name or "Unknown")
                if ok:
                    summary["synced"] += 1
                else:
                    summary["failed"] += 1
            except Exception as e:
                self._set_incremental_sync_error(f"mongo_profile_sync_failed :: {profile_id} :: {e}")
                summary["failed"] += 1
        return summary

    def _flush_incremental_sync_batch(
        self,
        pending_profiles: List[Dict[str, Any]],
        *,
        is_final_flush: bool = False,
    ) -> None:
        if not self.incremental_sync_enabled:
            return
        if not pending_profiles:
            return

        self._incremental_sync_stats["batches_attempted"] += 1
        print(
            "[SyncBatch] Flushing "
            f"{len(pending_profiles)} profiles"
            f"{' (final)' if is_final_flush else ''}..."
        )

        profile_name_map: Dict[str, str] = {}
        sync_chunks: List[Dict[str, Any]] = []
        for entry in pending_profiles:
            profile_id = str(entry.get("profile_id") or "").strip()
            if not profile_id:
                continue
            profile_name = str(entry.get("profile_name") or "Unknown")
            profile_name_map[profile_id] = profile_name
            chunks_path = entry.get("chunks_path")
            loaded = self._load_sync_chunks_for_profile(
                profile_id=profile_id,
                profile_name=profile_name,
                chunks_path=chunks_path,
            )
            sync_chunks.extend(loaded)

        self._incremental_sync_stats["profiles_enqueued"] += len(profile_name_map)
        self._incremental_sync_stats["chunks_loaded"] += len(sync_chunks)

        uploaded_profile_ids: Set[str] = set(profile_name_map.keys())
        pinecone_summary = self._upload_chunks_batch_incremental(sync_chunks)
        self._incremental_sync_stats["vectors_uploaded"] += int(pinecone_summary.get("uploaded_vectors", 0))
        self._incremental_sync_stats["vectors_failed"] += int(pinecone_summary.get("failed_vectors", 0))
        if not self.incremental_skip_pinecone:
            uploaded_profile_ids = set(pinecone_summary.get("uploaded_profile_ids") or set())

        mongo_targets = profile_name_map
        if not self.incremental_skip_pinecone and uploaded_profile_ids:
            mongo_targets = {
                profile_id: profile_name_map.get(profile_id, "Unknown")
                for profile_id in uploaded_profile_ids
            }
        elif not self.incremental_skip_pinecone and not uploaded_profile_ids:
            mongo_targets = {}

        mongo_summary = self._sync_profiles_batch_incremental(mongo_targets)
        self._incremental_sync_stats["profiles_synced_mongo"] += int(mongo_summary.get("synced", 0))
        self._incremental_sync_stats["profiles_failed_mongo"] += int(mongo_summary.get("failed", 0))
        self._incremental_sync_stats["batches_completed"] += 1

        print(
            "[SyncBatch] Done | "
            f"chunks={len(sync_chunks)}, "
            f"vectors_uploaded={pinecone_summary.get('uploaded_vectors', 0)}, "
            f"vectors_failed={pinecone_summary.get('failed_vectors', 0)}, "
            f"mongo_synced={mongo_summary.get('synced', 0)}, "
            f"mongo_failed={mongo_summary.get('failed', 0)}"
        )

    def _get_incremental_sync_stats_snapshot(self) -> Dict[str, Any]:
        return dict(self._incremental_sync_stats)
    
    async def scrape_profile(self, url: str, row_data: Dict) -> Optional[Dict]:
        """
        Scrape a single profile URL
        
        Args:
            url: Profile URL to scrape
            row_data: Additional data from Excel row
            
        Returns:
            Scraped profile data or None if failed
        """
        try:
            result = await self.scraper.extract_all(str(url))
            
            # Combine all text content + build source list
            combined_text_parts = []
            combined_headings = []
            combined_paragraphs = []
            source_items: List[Dict[str, Any]] = []

            # Get profile data early (needed for relevance filtering)
            profile_data = result.get('profile_data', {})
            # Prioritize name from Excel sheet (more accurate) over scraped name
            profile_name = row_data.get('name', '') or profile_data.get('name', '') or 'Unknown'
            # Clean up the name (remove extra whitespace)
            if profile_name:
                profile_name = ' '.join(profile_name.split())

            # Prefer source_records if provided
            source_records = result.get("source_records") or []
            if source_records:
                for rec in source_records:
                    source_type = rec.get("source_type", "webpage")
                    source_url = rec.get("source_url") or rec.get("resolved_url") or ""
                    resolved_url = rec.get("resolved_url") or source_url
                    content_text = rec.get("content") or ""
                    fetch_meta = rec.get("fetch_metadata") or {}
                    status = rec.get("processing_status", "success")

                    if not content_text and source_type in ("cv",) and status in ["binary_document", "failed", "auth_required"]:
                        cv_url = resolved_url or source_url
                        if cv_url:
                            print(f"  📄 Extracting CV text from binary document: {cv_url[:60]}...")
                            cv_text = self.extract_cv_text(cv_url)
                            if cv_text and self._is_reasonable_text(cv_text):
                                content_text = cv_text
                                print(f"  Extracted {len(cv_text)} characters from CV")
                            else:
                                print("  Warning: failed to extract text from CV")

                    if not content_text:
                        continue
                    if self.source_quality_filter_enabled and not self._is_reasonable_text(content_text):
                        allow_low_quality = self._allow_official_profile_page_despite_low_quality(
                            content_text=content_text,
                            source_type=source_type,
                            source_url=resolved_url or source_url,
                            profile_url=str(url),
                            profile_name=profile_name,
                        )
                        if not allow_low_quality:
                            print(f"  [Filter] Skipping low-quality content from {source_url[:60]}...")
                            continue
                    if self.profile_relevance_filter_enabled and not self._source_matches_target_profile(
                        source_type=source_type,
                        source_url=resolved_url or source_url,
                        link_text=rec.get("title") or rec.get("link_text") or "",
                        content_text=content_text,
                        profile_name=profile_name,
                    ):
                        print(f"  [Filter] Skipping content not specific to target profile from {source_url[:60]}...")
                        continue

                    source_items.append({
                        "source_type": source_type,
                        "source_url": source_url,
                        "resolved_url": resolved_url,
                        "content_text": content_text,
                        "fetch_metadata": fetch_meta,
                        "link_text": rec.get("title") or rec.get("link_text") or "",
                    })
            else:
                # Legacy fallback using profile page + body links
                profile_text = result.get('text_content', {})
                if profile_text.get('full_text'):
                    source_items.append({
                        "source_type": "profile_page",
                        "source_url": str(url),
                        "resolved_url": str(url),
                        "content_text": profile_text.get('full_text', ''),
                        "fetch_metadata": {},
                        "link_text": "Profile Page",
                    })
                    combined_headings.extend(profile_text.get('headings', []))
                    combined_paragraphs.extend(profile_text.get('paragraphs', []))

                body_links = result.get('body_links_with_content', [])
                if self.debug_links and body_links:
                    print("  [ScraperDebug] Body links with content:")
                    for link in body_links:
                        dbg_url = link.get('source_url') or link.get('url') or link.get('resolved_url')
                        print(
                            f"    - category={link.get('category')} "
                            f"is_document={link.get('is_document')} "
                            f"status={link.get('processing_status')} "
                            f"url={dbg_url}"
                        )

                for link in body_links:
                    content_text = link.get('content', '')
                    if content_text and link.get('processing_status') == 'success':
                        if self.source_quality_filter_enabled and not self._is_reasonable_text(content_text):
                            allow_low_quality = self._allow_official_profile_page_despite_low_quality(
                                content_text=content_text,
                                source_type=link.get("category") or "webpage",
                                source_url=link.get('resolved_url') or link.get('source_url') or link.get('url') or '',
                                profile_url=str(url),
                                profile_name=profile_name,
                            )
                            if not allow_low_quality:
                                continue
                        if self.profile_relevance_filter_enabled and not self._source_matches_target_profile(
                            source_type=link.get("category") or "webpage",
                            source_url=link.get('resolved_url') or link.get('source_url') or link.get('url') or '',
                            link_text=link.get("link_text") or "",
                            content_text=content_text,
                            profile_name=profile_name,
                        ):
                            continue
                        source_items.append({
                            "source_type": link.get("category") or "webpage",
                            "source_url": link.get('source_url') or link.get('url') or '',
                            "resolved_url": link.get('resolved_url') or link.get('source_url') or link.get('url') or '',
                            "content_text": content_text,
                            "fetch_metadata": link.get("fetch_metadata") or {},
                            "link_text": link.get("link_text") or "",
                        })

            # Build combined text (legacy)
            for item in source_items:
                source_type = item.get("source_type", "webpage")
                link_text = item.get("link_text") or extract_title_from_text(item.get("content_text", ""))
                source_url = item.get("source_url", "")
                content_text = item.get("content_text", "")
                if source_type == "profile_page":
                    combined_text_parts.append(f"=== PROFILE PAGE ===\n{content_text}")
                else:
                    combined_text_parts.append(
                        f"\n\n=== {source_type.upper()}: {link_text} ({source_url}) ===\n{content_text}"
                    )

            combined_text = '\n\n'.join(combined_text_parts) if combined_text_parts else ''

            # Collect all URLs
            all_urls = [str(url)]
            for item in source_items:
                link_url = item.get("source_url") or item.get("resolved_url")
                if link_url and link_url not in all_urls:
                    all_urls.append(link_url)
            
            return {
                'profile_data': profile_data,
                'profile_name': profile_name,
                'profile_url': str(url),
                'all_urls': all_urls,
                'combined_text': combined_text,
                'source_items': source_items,
                'combined_headings': combined_headings,
                'combined_paragraphs': combined_paragraphs,
                'row_data': row_data,
                'extraction_metadata': result.get('extraction_metadata', {})
            }
            
        except Exception as e:
            print(f"  Error scraping {url}: {str(e)}")
            self._log_error(f"[scrape_profile] {url} :: {e}\n{traceback.format_exc()}")
            return None

    async def scrape_profile_urls(self, urls: List[str], row_data: Dict) -> Optional[Dict]:
        """
        Scrape multiple URLs and merge into a single profile payload.

        Args:
            urls: List of seed URLs for a single profile
            row_data: Additional data (e.g., name)

        Returns:
            Combined scraped data or None if all failed
        """
        combined_text_parts = []
        combined_headings = []
        combined_paragraphs = []
        all_urls_lists: List[List[str]] = []
        profile_names: List[str] = []
        all_source_items: List[Dict[str, Any]] = []
        successful = 0
        failed = 0

        total = len(urls)
        for idx, url in enumerate(urls, 1):
            print(f"  [Seed {idx}/{total}] Scraping: {url[:60]}...")
            scraped = await self.scrape_profile(url, row_data)
            if not scraped or not scraped.get("combined_text", "").strip():
                failed += 1
                continue

            successful += 1
            profile_names.append(scraped.get("profile_name", ""))
            all_urls_lists.append(scraped.get("all_urls", []))
            combined_headings.extend(scraped.get("combined_headings", []))
            combined_paragraphs.extend(scraped.get("combined_paragraphs", []))
            all_source_items.extend(scraped.get("source_items", []))

            combined_text_parts.append(
                f"=== SEED URL {idx}: {scraped.get('profile_url', url)} ===\n{scraped.get('combined_text', '')}"
            )

        if not combined_text_parts:
            print("  All seed URLs failed or returned no content")
            self._log_error("[scrape_profile_urls] All seed URLs failed or returned no content")
            return None

        # Resolve profile name preference: explicit row_data wins.
        # Scraped per-page titles are unreliable (HTML <title> often returns
        # things like "beta.centralseminary.edu", "Shopping Basket", "Legacy"
        # or a finding-aid collection name), so we only fall back to scraped
        # candidates after stripping obvious junk.
        profile_name = row_data.get("name", "").strip()
        if not profile_name:
            for name in profile_names:
                cand = (name or "").strip()
                if not cand:
                    continue
                if cand.lower() in {"unknown", "shopping basket", "legacy", "loading"}:
                    continue
                if "." in cand and " " not in cand:
                    # Looks like a domain ("beta.centralseminary.edu").
                    continue
                if cand.lower().endswith(("papers", "collection", "archive", "archives")):
                    # Finding-aid collection title, not a person's name.
                    continue
                profile_name = cand
                break
        if not profile_name:
            profile_name = "Unknown"

        combined_text = "\n\n".join(combined_text_parts)
        merged_urls = merge_url_lists(all_urls_lists, seed_urls=urls)

        return {
            "profile_name": profile_name,
            "combined_text": combined_text,
            "combined_headings": combined_headings,
            "combined_paragraphs": combined_paragraphs,
            "all_urls": merged_urls,
            "source_items": all_source_items,
            "seed_urls_total": total,
            "seed_urls_successful": successful,
            "seed_urls_failed": failed,
        }
    
    def extract_cv_text(self, cv_url: str) -> str:
        """
        Extract text from a CV document URL using DocumentProcessor
        
        Args:
            cv_url: URL to the CV document
            
        Returns:
            Extracted text or empty string if failed
        """
        try:
            doc_result = self.document_processor.process_document(cv_url)
            content = doc_result.get('content', '')
            if content and content.strip():
                return content.strip()
            # Fallback: direct document extraction is more reliable for many CV PDFs.
            import io
            import ssl
            import urllib.request

            req = urllib.request.Request(cv_url, headers={"User-Agent": "Mozilla/5.0"})
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            with urllib.request.urlopen(req, context=ctx, timeout=45) as resp:
                data = resp.read()
                content_type = (resp.headers.get("Content-Type") or "").lower()

            if not data:
                return ""

            if cv_url.lower().endswith(".docx") or "officedocument" in content_type:
                try:
                    from docx import Document  # type: ignore

                    doc = Document(io.BytesIO(data))
                    parts = [p.text.strip() for p in doc.paragraphs if (p.text or "").strip()]
                    text = "\n".join(parts).strip()
                    if text:
                        return text
                except Exception:
                    pass

            try:
                import fitz  # type: ignore

                doc = fitz.open(stream=data, filetype="pdf")
                parts = []
                for page in doc:
                    try:
                        page_text = page.get_text("text", sort=True) or ""
                    except Exception:
                        page_text = ""
                    if page_text.strip():
                        parts.append(page_text)
                text = "\n".join(part.strip() for part in parts if part and part.strip()).strip()
                if text:
                    return text
            except Exception:
                pass
            return ""
        except Exception as e:
            print(f"  Warning: error extracting CV text from {cv_url[:60]}...: {str(e)}")
            return ""
    
    def clean_text(self, text: str, profile_url: str) -> str:
        """
        Clean extracted text using data cleaning service
        
        Args:
            text: Raw text to clean
            profile_url: Profile URL for context
            
        Returns:
            Cleaned text
        """
        if not text or not text.strip():
            return ""
        
        try:
            # Use cleaning service to clean and chunk
            chunks = self.cleaning_service.clean_and_chunk_text(
                text=text,
                profile_url=profile_url,
                section_header=""
            )
            
            # Combine cleaned chunks back into text
            cleaned_text = " ".join([chunk.get("text", "") for chunk in chunks]).strip()
            return cleaned_text
            
        except Exception as e:
            print(f"  Warning: cleaning failed: {str(e)}, using raw text")
            return text
    
    def create_chunked_profile(
        self,
        profile_id: str,
        cleaned_text: str,
        raw_text: Optional[str] = None,
        *,
        source_chunks: Optional[List[Dict[str, Any]]] = None,
        sources_meta: Optional[List[Dict[str, Any]]] = None,
        profile_name: Optional[str] = None,
    ) -> Optional[Path]:
        """
        Create section-aware chunks for the profile.

        When ``source_chunks`` is supplied (the per-source chunks produced by
        :meth:`_process_sources_with_intent_gates`), we now build the final
        ``chunks.json`` directly from them — preserving source_id,
        offset_start/offset_end, language, allowed_use, and quote_ok — and
        section labels are derived from each source's URL/link_text rather
        than from an LLM that re-segmented the concatenated blob (which used
        to produce ghost sections like "Quaker Restrictions" or
        "Byrd's Profile" that did not actually describe the subject).

        ``cleaned_text`` is retained only for the legacy fallback path used
        when no per-source chunks are available.
        """
        if not self.use_llm_chunking or not self.chunking_pipeline:
            return None

        try:
            if source_chunks:
                output_path = self.chunking_pipeline.build_chunks_from_source_chunks(
                    profile_id=profile_id,
                    source_chunks=source_chunks,
                    sources_meta=sources_meta or [],
                    profile_name=profile_name or "",
                    raw_text=raw_text,
                    cleaned_text=cleaned_text,
                )
                return output_path

            if not cleaned_text or not cleaned_text.strip():
                return None

            output_path = self.chunking_pipeline.process_profile(
                profile_id,
                cleaned_text,
                raw_text=raw_text
            )
            return output_path
        except Exception as e:
            print(f"  Warning: chunking failed: {str(e)}")
            return None

    def _minimal_profile_excerpt(
        self,
        text: str,
        *,
        profile_name: str,
        email: str = "",
        max_chars: int = 420,
    ) -> str:
        flat_text = self._normalize_inline_whitespace(text)
        if not flat_text:
            return ""

        name_lc = (profile_name or "").strip().lower()
        email_lc = (email or "").strip().lower()
        last_name = ""
        if name_lc:
            parts = [part for part in re.findall(r"[a-z0-9]+", name_lc) if len(part) >= 2]
            if parts:
                last_name = parts[-1]

        kept_lines: List[str] = []
        for raw_line in (text or "").splitlines():
            line = self._normalize_inline_whitespace(raw_line)
            if len(line) < 4:
                continue
            line_lc = line.lower()
            if any(token in line_lc for token in ("skip to main content", "cookie", "privacy", "terms of use")):
                continue
            if (
                (name_lc and name_lc in line_lc)
                or (last_name and last_name in line_lc)
                or (email_lc and email_lc in line_lc)
                or any(token in line_lc for token in ("professor", "advisor", "department", "research", "email", "phone"))
            ):
                kept_lines.append(line)
            if len(" ".join(kept_lines)) >= max_chars:
                break

        excerpt = " ".join(kept_lines) if kept_lines else flat_text[:max_chars]
        return excerpt[:max_chars].strip()

    def _build_profile_fallback_chunk_inputs(
        self,
        *,
        profile_id: str,
        profile_name: str,
        profile_url: str,
        scraped_data: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        source_items = scraped_data.get("source_items") or []
        profile_data = scraped_data.get("profile_data") or {}

        official_item = None
        for item in source_items:
            if (item.get("source_type") or "") == "profile_page":
                official_item = item
                break

        source_url = (
            (official_item or {}).get("resolved_url")
            or (official_item or {}).get("source_url")
            or profile_url
        )
        if not source_url:
            return None

        role_value = self._normalize_inline_whitespace(
            str(profile_data.get("position") or profile_data.get("title") or "")
        )
        department_value = self._normalize_inline_whitespace(str(profile_data.get("department") or ""))
        email_value = self._normalize_inline_whitespace(str(profile_data.get("email") or ""))
        excerpt_text = self._minimal_profile_excerpt(
            (official_item or {}).get("content_text") or str(profile_data.get("full_text") or ""),
            profile_name=profile_name,
            email=email_value,
        )

        fact_lines: List[str] = []
        if profile_name:
            fact_lines.append(f"Name: {profile_name}")
        if role_value:
            fact_lines.append(f"Role: {role_value}")
        if department_value:
            fact_lines.append(f"Department: {department_value}")
        if email_value:
            fact_lines.append(f"Email: {email_value}")
        if source_url:
            fact_lines.append(f"Official profile URL: {source_url}")
        if excerpt_text:
            fact_lines.append(f"Profile excerpt: {excerpt_text}")

        chunk_text = "\n".join(fact_lines).strip()
        if len(chunk_text) < 40:
            return None

        source_id = make_source_id(source_url)
        source_meta = {
            "source_id": source_id,
            "source_type": "profile_page",
            "source_url": source_url,
            "resolved_url": source_url,
            "link_text": (official_item or {}).get("link_text") or "Official Profile Facts",
            "allowed_use": "facts_only",
            "language": detect_language(chunk_text),
            "fetch_metadata": (official_item or {}).get("fetch_metadata") or {},
        }
        source_chunk = {
            "chunk_id": f"{source_id}:fallback_profile_fact_chunk",
            "source_id": source_id,
            "text": chunk_text,
            "raw_text": chunk_text,
            "offset_start": 0,
            "offset_end": len(chunk_text),
            "allowed_use": "facts_only",
            "language": source_meta["language"],
            "subject_mention": True,
            "text_hash": compute_text_hash(chunk_text),
        }
        source_item = {
            "source_type": "profile_page",
            "source_url": source_url,
            "resolved_url": source_url,
            "content_text": chunk_text,
            "fetch_metadata": source_meta["fetch_metadata"],
            "link_text": source_meta["link_text"],
        }
        return {
            "source_chunk": source_chunk,
            "source_meta": source_meta,
            "source_item": source_item,
        }

    def _ensure_profile_chunk_fallback(
        self,
        *,
        profile_id: str,
        profile_name: str,
        profile_url: str,
        scraped_data: Dict[str, Any],
        source_chunks: List[Dict[str, Any]],
        sources_meta: List[Dict[str, Any]],
        allowed_source_items: List[Dict[str, Any]],
        intent_summary: Dict[str, Any],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
        if source_chunks:
            return source_chunks, sources_meta, allowed_source_items, intent_summary

        fallback = self._build_profile_fallback_chunk_inputs(
            profile_id=profile_id,
            profile_name=profile_name,
            profile_url=profile_url,
            scraped_data=scraped_data,
        )
        if not fallback:
            return source_chunks, sources_meta, allowed_source_items, intent_summary

        updated_source_chunks = list(source_chunks) + [fallback["source_chunk"]]
        updated_sources_meta = list(sources_meta)
        if not any((item.get("source_id") or "") == fallback["source_meta"]["source_id"] for item in updated_sources_meta):
            updated_sources_meta.append(fallback["source_meta"])
        updated_allowed_items = list(allowed_source_items)
        if not updated_allowed_items:
            updated_allowed_items.append(fallback["source_item"])
        updated_intent_summary = dict(intent_summary or {})
        updated_intent_summary["fallback_chunk_used"] = True
        updated_intent_summary["accepted_chunk_count"] = len(updated_source_chunks)
        updated_intent_summary["accepted_source_count"] = max(
            int(updated_intent_summary.get("accepted_source_count", 0)),
            len(updated_allowed_items),
        )
        return (
            updated_source_chunks,
            updated_sources_meta,
            updated_allowed_items,
            updated_intent_summary,
        )

    def save_profile_json(
        self,
        profile_id: str,
        profile_name: str,
        profile_url: str,
        all_urls: List[str],
        combined_text: str,
        cleaned_text: str,
        chunks_path: Optional[Path] = None,
        source_chunks_path: Optional[Path] = None,
        claims_path: Optional[Path] = None,
        source_registry_path: Optional[str] = None,
        intent_summary: Optional[Dict[str, Any]] = None
    ) -> Path:
        """
        Save profile data to JSON file
        
        Args:
            profile_id: Profile ID
            profile_name: Profile name
            profile_url: Profile URL
            all_urls: List of all URLs
            combined_text: Raw combined text
            cleaned_text: Cleaned text
            chunks_path: Path to chunks.json file if available
            
        Returns:
            Path to saved JSON file
        """
        # Create profile directory
        profile_dir = self.output_dir / "profiles" / profile_id
        profile_dir.mkdir(parents=True, exist_ok=True)
        
        # Load chunks if available
        chunks_data = None
        if chunks_path and chunks_path.exists():
            try:
                with open(chunks_path, 'r', encoding='utf-8') as f:
                    chunks_data = json.load(f)
            except Exception as e:
                print(f"  Warning: failed to load chunks: {str(e)}")
        
        # Create profile JSON
        profile_json = {
            "profile_id": profile_id,
            "name": profile_name,
            "profile_url": profile_url,
            "all_urls": all_urls,
            "raw_text": combined_text,
            "clean_text": cleaned_text,
            "has_cv": bool(cleaned_text and len(cleaned_text) > 100),
            "chunks_available": chunks_data is not None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Add chunks reference if available
        if chunks_data:
            profile_json["chunks_file"] = str(chunks_path.relative_to(self.output_dir))
        if source_chunks_path:
            profile_json["source_chunks_file"] = str(source_chunks_path.relative_to(self.output_dir))
        if claims_path:
            profile_json["claims_file"] = str(claims_path.relative_to(self.output_dir))
        if source_registry_path:
            try:
                profile_json["source_registry"] = str(Path(source_registry_path))
            except Exception:
                profile_json["source_registry"] = source_registry_path
        if intent_summary is not None:
            profile_json["intent_gates"] = intent_summary

        # Save JSON file
        json_path = profile_dir / f"{profile_id}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(profile_json, f, indent=2, ensure_ascii=False)
        
        return json_path

    @staticmethod
    def _stable_profile_id_from_row(row_data: Optional[Dict[str, Any]]) -> Optional[str]:
        if not isinstance(row_data, dict):
            return None
        for key in ("Scholar Profile ID", "profile_id", "Profile ID"):
            value = row_data.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text and text.lower() != "nan":
                return text
        return None
    
    async def process_single_profile(
        self,
        idx: int,
        total: int,
        url: str,
        row_data: Dict
    ) -> Dict[str, Any]:
        """
        Process a single profile through the complete pipeline
        
        Args:
            idx: Current index
            total: Total number of profiles
            url: Profile URL
            row_data: Excel row data
            
        Returns:
            Processing result dictionary
        """
        print(f"\n[{idx}/{total}] Processing: {url[:60]}...")
        
        # Step 1: Scrape
        scraped_data = await self.scrape_profile(url, row_data)
        if not scraped_data:
            self._log_error(f"[process_single_profile] {url} :: Scraping failed")
            return {
                'url': url,
                'status': 'failed',
                'error': 'Scraping failed',
                'profile_id': None
            }
        
        profile_name = scraped_data['profile_name']
        combined_text = scraped_data['combined_text']
        
        if not combined_text or not combined_text.strip():
            self._log_error(f"[process_single_profile] {url} :: No content extracted")
            return {
                'url': url,
                'status': 'failed',
                'error': 'No content extracted',
                'profile_id': None
            }
        
        # Use a stable ID from the source workbook when available so retries
        # update the same scholar instead of forking duplicates.
        profile_id = self._stable_profile_id_from_row(row_data) or str(uuid.uuid4())

        # Step 2: Register sources + per-source chunking
        source_items = scraped_data.get("source_items", [])
        if not source_items:
            self._log_error(f"[process_single_profile] {url} :: No source items after scrape")
            return {
                'url': url,
                'status': 'failed',
                'error': 'No source items extracted',
                'profile_id': None
            }

        processed_sources = self._process_sources_with_intent_gates(
            source_items=source_items,
            profile_name=profile_name,
            profile_id=profile_id,
            profile_url=url,
            search_meta=row_data.get("_search_meta") if isinstance(row_data, dict) else None,
        )
        source_chunks: List[Dict[str, Any]] = processed_sources["source_chunks"]
        sources_meta: List[Dict[str, Any]] = processed_sources["sources_meta"]
        allowed_source_items: List[Dict[str, Any]] = processed_sources["allowed_source_items"]
        intent_summary: Dict[str, Any] = processed_sources.get("intent_summary", {})
        profile_gate_pass = bool(processed_sources.get("profile_gate_pass", True))
        profile_gate_reason = processed_sources.get("profile_gate_reason", "")
        source_chunks, sources_meta, allowed_source_items, intent_summary = self._ensure_profile_chunk_fallback(
            profile_id=profile_id,
            profile_name=profile_name,
            profile_url=url,
            scraped_data=scraped_data,
            source_chunks=source_chunks,
            sources_meta=sources_meta,
            allowed_source_items=allowed_source_items,
            intent_summary=intent_summary,
        )

        print(
            "  [Intent] "
            f"sources={intent_summary.get('accepted_source_count', 0)}/{len(source_items)} | "
            f"chunks={intent_summary.get('accepted_chunk_count', 0)} | "
            f"covered_required={intent_summary.get('required_types_covered_count', 0)}"
        )

        # Save per-source chunks + claims
        source_chunks_path = self._save_source_chunks(
            profile_id,
            source_chunks,
            sources_meta,
            intent_summary=intent_summary,
        )
        claims = extract_claims(
            profile_id,
            source_chunks,
            profile_name=profile_name,
            sources_meta=sources_meta,
        )
        claims_path = self._save_claims(profile_id, claims)

        if not allowed_source_items:
            reason = profile_gate_reason or "no_sources_passed_policy_or_intent_gates"
            self._log_error(f"[process_single_profile] {url} :: {reason}")
            return {
                'url': url,
                'status': 'failed',
                'error': reason,
                'profile_id': None,
                'intent_summary': intent_summary,
            }

        if self.intent_gating_enabled and not profile_gate_pass:
            reason = profile_gate_reason or "profile_intent_gate_failed"
            self._log_error(f"[process_single_profile] {url} :: {reason}")
            return {
                'url': url,
                'status': 'failed',
                'error': reason,
                'profile_id': None,
                'intent_summary': intent_summary,
            }

        # Rebuild combined_text using allowed sources (legacy output)
        filtered_urls = []
        seen_urls = set()
        for item in allowed_source_items:
            u = item.get("source_url") or item.get("resolved_url")
            if not u:
                continue
            u_norm = normalize_url(u)
            if u_norm in seen_urls:
                continue
            seen_urls.add(u_norm)
            filtered_urls.append(u)

        combined_text_parts = []
        for item in allowed_source_items:
            source_type = item.get("source_type", "webpage")
            link_text = item.get("link_text") or extract_title_from_text(item.get("content_text", ""))
            source_url = item.get("source_url", "")
            content_text = item.get("content_text", "")
            if source_type == "profile_page":
                combined_text_parts.append(f"=== PROFILE PAGE ===\n{content_text}")
            else:
                combined_text_parts.append(
                    f"\n\n=== {source_type.upper()}: {link_text} ({source_url}) ===\n{content_text}"
                )
        if combined_text_parts:
            combined_text = "\n\n".join(combined_text_parts)
        else:
            combined_text = ""

        # Step 3: Clean combined text. Skip the LLM cleaner when we
        # have per-source chunks — it rewrites text and breaks the
        # raw_text → chunk offset mapping for the chatbot.
        combined_text = self._filter_text_for_profile(combined_text, profile_name)
        if self.llm_cleaner and not source_chunks:
            try:
                combined_text = self.llm_cleaner.clean_text(combined_text, use_chunking=True)
            except Exception as e:
                self._log_error(f"[llm_cleaning] {url} :: {e}\n{traceback.format_exc()}")
        print(f"  [Cleaning] Cleaning text ({len(combined_text)} chars)...")
        cleaned_text = self.clean_text(combined_text, url)
        print(f"  [Cleaning] Cleaned to {len(cleaned_text)} chars")
        
        # Step 3: Create chunks (if enabled)
        chunks_path = None
        if self.use_llm_chunking and (cleaned_text or source_chunks):
            print(f"  [Chunking] Creating section-aware chunks...")
            chunks_path = self.create_chunked_profile(
                profile_id,
                cleaned_text,
                raw_text=combined_text,
                source_chunks=source_chunks,
                sources_meta=sources_meta,
                profile_name=profile_name,
            )
            if chunks_path:
                print(f"  [Chunking] Saved chunks to {chunks_path}")
            else:
                print("  [Chunking] Warning: chunking skipped or failed")

        # Step 4: Save JSON
        print(f"  [Saving] Saving profile JSON...")
        json_path = self.save_profile_json(
            profile_id=profile_id,
            profile_name=profile_name,
            profile_url=url,
            all_urls=filtered_urls or scraped_data['all_urls'],
            combined_text=combined_text,
            cleaned_text=cleaned_text,
            chunks_path=chunks_path,
            source_chunks_path=source_chunks_path,
            claims_path=claims_path,
            source_registry_path=self.source_registry_path,
            intent_summary=intent_summary,
        )
        print(f"  [Saving] Saved to {json_path}")
        
        return {
            'url': url,
            'status': 'success',
            'profile_id': profile_id,
            'profile_name': profile_name,
            'chunks_path': str(chunks_path) if chunks_path else None,
            'source_chunks_path': str(source_chunks_path) if source_chunks_path else None,
            'claims_path': str(claims_path) if claims_path else None,
            'json_path': str(json_path),
            'intent_summary': intent_summary,
        }

    async def process_profile_from_urls(
        self,
        urls: List[str],
        profile_url: Optional[str] = None,
        profile_name: Optional[str] = None,
        profile_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Process a single profile from a list of seed URLs.
        """
        primary_url = profile_url or (urls[0] if urls else "")
        print(f"\n[1/1] Processing URL list ({len(urls)} URLs)")
        if not primary_url:
            return {
                "url": "",
                "status": "failed",
                "error": "No URLs provided",
                "profile_id": None,
            }

        row_data = {}
        if profile_name:
            row_data["name"] = profile_name

        # Step 1: Scrape all seed URLs
        scraped_data = await self.scrape_profile_urls(urls, row_data)
        if not scraped_data:
            self._log_error(f"[process_profile_from_urls] {primary_url} :: Scraping failed for all seed URLs")
            return {
                "url": primary_url,
                "status": "failed",
                "error": "Scraping failed for all seed URLs",
                "profile_id": None,
            }

        # Use resolved profile name
        resolved_name = profile_name or scraped_data.get("profile_name", "Unknown")
        combined_text = scraped_data["combined_text"]

        if not combined_text or not combined_text.strip():
            self._log_error(f"[process_profile_from_urls] {primary_url} :: No content extracted")
            return {
                "url": primary_url,
                "status": "failed",
                "error": "No content extracted",
                "profile_id": None,
            }

        # When URL-list mode is driven by a known scholar record, preserve its
        # stable ID so Pinecone/Mongo/SQLite all converge on one document.
        profile_id = str(profile_id).strip() if profile_id else str(uuid.uuid4())

        # Step 2: Register sources + per-source chunking
        source_items = scraped_data.get("source_items", [])
        if not source_items:
            self._log_error(f"[process_profile_from_urls] {primary_url} :: No source items after scrape")
            return {
                "url": primary_url,
                "status": "failed",
                "error": "No source items extracted",
                "profile_id": None,
            }

        processed_sources = self._process_sources_with_intent_gates(
            source_items=source_items,
            profile_name=resolved_name,
            profile_id=profile_id,
            profile_url=primary_url,
            search_meta=None,
        )
        source_chunks: List[Dict[str, Any]] = processed_sources["source_chunks"]
        sources_meta: List[Dict[str, Any]] = processed_sources["sources_meta"]
        allowed_source_items: List[Dict[str, Any]] = processed_sources["allowed_source_items"]
        intent_summary: Dict[str, Any] = processed_sources.get("intent_summary", {})
        profile_gate_pass = bool(processed_sources.get("profile_gate_pass", True))
        profile_gate_reason = processed_sources.get("profile_gate_reason", "")
        source_chunks, sources_meta, allowed_source_items, intent_summary = self._ensure_profile_chunk_fallback(
            profile_id=profile_id,
            profile_name=resolved_name,
            profile_url=primary_url,
            scraped_data=scraped_data,
            source_chunks=source_chunks,
            sources_meta=sources_meta,
            allowed_source_items=allowed_source_items,
            intent_summary=intent_summary,
        )

        print(
            "  [Intent] "
            f"sources={intent_summary.get('accepted_source_count', 0)}/{len(source_items)} | "
            f"chunks={intent_summary.get('accepted_chunk_count', 0)} | "
            f"covered_required={intent_summary.get('required_types_covered_count', 0)}"
        )

        # Save per-source chunks + claims
        source_chunks_path = self._save_source_chunks(
            profile_id,
            source_chunks,
            sources_meta,
            intent_summary=intent_summary,
        )
        claims = extract_claims(
            profile_id,
            source_chunks,
            profile_name=resolved_name,
            sources_meta=sources_meta,
        )
        claims_path = self._save_claims(profile_id, claims)

        if not allowed_source_items:
            reason = profile_gate_reason or "no_sources_passed_policy_or_intent_gates"
            self._log_error(f"[process_profile_from_urls] {primary_url} :: {reason}")
            return {
                "url": primary_url,
                "status": "failed",
                "error": reason,
                "profile_id": None,
                "intent_summary": intent_summary,
            }

        if self.intent_gating_enabled and not profile_gate_pass:
            reason = profile_gate_reason or "profile_intent_gate_failed"
            self._log_error(f"[process_profile_from_urls] {primary_url} :: {reason}")
            return {
                "url": primary_url,
                "status": "failed",
                "error": reason,
                "profile_id": None,
                "intent_summary": intent_summary,
            }

        # Rebuild combined_text using allowed sources (legacy output)
        filtered_urls = []
        seen_urls = set()
        for item in allowed_source_items:
            u = item.get("source_url") or item.get("resolved_url")
            if not u:
                continue
            u_norm = normalize_url(u)
            if u_norm in seen_urls:
                continue
            seen_urls.add(u_norm)
            filtered_urls.append(u)

        combined_text_parts = []
        for item in allowed_source_items:
            source_type = item.get("source_type", "webpage")
            link_text = item.get("link_text") or extract_title_from_text(item.get("content_text", ""))
            source_url = item.get("source_url", "")
            content_text = item.get("content_text", "")
            if source_type == "profile_page":
                combined_text_parts.append(f"=== PROFILE PAGE ===\n{content_text}")
            else:
                combined_text_parts.append(
                    f"\n\n=== {source_type.upper()}: {link_text} ({source_url}) ===\n{content_text}"
                )
        if combined_text_parts:
            combined_text = "\n\n".join(combined_text_parts)
        else:
            combined_text = ""

        # Step 3: Clean combined text. We deliberately *skip* the LLM
        # cleaner here when we have per-source chunks — the LLM rewrites
        # text and breaks raw_text-to-chunk offset mapping. The legacy
        # combined cleaning path is only useful when source_chunks is
        # empty (the fallback chunker further down still benefits).
        combined_text = self._filter_text_for_profile(combined_text, resolved_name)
        if self.llm_cleaner and not source_chunks:
            try:
                combined_text = self.llm_cleaner.clean_text(combined_text, use_chunking=True)
            except Exception as e:
                self._log_error(f"[llm_cleaning] {primary_url} :: {e}\n{traceback.format_exc()}")
        print(f"  [Cleaning] Cleaning text ({len(combined_text)} chars)...")
        cleaned_text = self.clean_text(combined_text, primary_url)
        print(f"  [Cleaning] Cleaned to {len(cleaned_text)} chars")

        # Step 3: Create chunks (if enabled)
        chunks_path = None
        if self.use_llm_chunking and (cleaned_text or source_chunks):
            print(f"  [Chunking] Creating section-aware chunks...")
            chunks_path = self.create_chunked_profile(
                profile_id,
                cleaned_text,
                raw_text=combined_text,
                source_chunks=source_chunks,
                sources_meta=sources_meta,
                profile_name=resolved_name,
            )
            if chunks_path:
                print(f"  [Chunking] Saved chunks to {chunks_path}")
            else:
                print("  [Chunking] Warning: chunking skipped or failed")

        # Step 4: Save JSON
        print(f"  [Saving] Saving profile JSON...")
        json_path = self.save_profile_json(
            profile_id=profile_id,
            profile_name=resolved_name,
            profile_url=primary_url,
            all_urls=filtered_urls or scraped_data["all_urls"],
            combined_text=combined_text,
            cleaned_text=cleaned_text,
            chunks_path=chunks_path,
            source_chunks_path=source_chunks_path,
            claims_path=claims_path,
            source_registry_path=self.source_registry_path,
            intent_summary=intent_summary,
        )
        print(f"  [Saving] Saved to {json_path}")

        return {
            "url": primary_url,
            "status": "success",
            "profile_id": profile_id,
            "profile_name": resolved_name,
            "chunks_path": str(chunks_path) if chunks_path else None,
            "source_chunks_path": str(source_chunks_path) if source_chunks_path else None,
            "claims_path": str(claims_path) if claims_path else None,
            "json_path": str(json_path),
            "intent_summary": intent_summary,
            "seed_urls_total": scraped_data.get("seed_urls_total", 0),
            "seed_urls_successful": scraped_data.get("seed_urls_successful", 0),
            "seed_urls_failed": scraped_data.get("seed_urls_failed", 0),
        }
    
    async def run(
        self,
        excel_path: str,
        limit: Optional[int] = None,
        start_from: int = 0
    ) -> Dict[str, Any]:
        """
        Run the complete pipeline
        
        Args:
            excel_path: Path to Excel file
            limit: Optional limit on number of profiles to process
            start_from: Start processing from this index (for resuming)
            
        Returns:
            Summary dictionary with results
        """
        print("="*80)
        print("UNIFIED PIPELINE: Excel → Scraping → Cleaning → Chunking → JSON")
        print("="*80)
        
        # Initialize scraper
        await self.initialize_scraper()
        
        # Step 1: Read Excel
        df = self.read_excel(excel_path)
        
        # Apply limit and start_from
        if start_from > 0:
            df = df.iloc[start_from:].reset_index(drop=True)
        if limit:
            df = df.head(limit)
        
        total = len(df)
        print(f"\n[Pipeline] Processing {total} profiles")
        print(f"[Pipeline] Output directory: {self.output_dir}")
        print(f"[Pipeline] Chunking output: {self.chunking_output_dir}")
        print(f"[Pipeline] LLM Chunking: {'Enabled' if self.use_llm_chunking else 'Disabled'}")
        
        # Process profiles
        results = []
        successful = 0
        ignored = 0
        failed = 0
        excel_tracker = self._open_excel_ignore_tracker(excel_path)
        pending_sync_profiles: List[Dict[str, Any]] = []
        
        for idx, (_, row) in enumerate(df.iterrows(), 1):
            url = row['profile_url']
            excel_row_number = None
            if "_excel_row_number" in row and pd.notna(row["_excel_row_number"]):
                try:
                    excel_row_number = int(row["_excel_row_number"])
                except Exception:
                    excel_row_number = None
            # Preserve all columns including 'name' for use in pipeline
            row_data = {col: str(row[col]) if pd.notna(row[col]) else '' 
                       for col in df.columns if col not in ('profile_url', '_excel_row_number')}
            # Ensure name column is available (case-insensitive)
            if 'name' not in row_data:
                # Try to find name column with different case
                name_cols = [col for col in df.columns if col.lower() == 'name']
                if name_cols:
                    row_data['name'] = str(row[name_cols[0]]) if pd.notna(row[name_cols[0]]) else ''

            if self.default_search_meta:
                row_data["_search_meta"] = self.default_search_meta

            try:
                result = await self.process_single_profile(idx, total, url, row_data)
                results.append(result)

                status = str(result.get("status") or "").strip().lower()
                if status not in {"success", "failed", "ignored"}:
                    status = "failed"
                if status == 'success':
                    successful += 1
                    if self.incremental_sync_enabled:
                        pending_sync_profiles.append(
                            {
                                "profile_id": result.get("profile_id"),
                                "profile_name": result.get("profile_name", "Unknown"),
                                "chunks_path": result.get("chunks_path"),
                            }
                        )
                        if len(pending_sync_profiles) >= self.incremental_sync_batch_size:
                            self._flush_incremental_sync_batch(
                                pending_sync_profiles,
                                is_final_flush=False,
                            )
                            pending_sync_profiles = []
                elif status == "ignored":
                    ignored += 1
                else:
                    failed += 1
                self._set_excel_ignore_status(
                    excel_tracker,
                    excel_row_number=excel_row_number,
                    ignored=(status == "ignored"),
                    reason=(result.get("ignore_reason", "") or result.get("error", "")) if status == "ignored" else "",
                    status=status,
                    scrape_issue=(result.get("error", "") if status == "failed" else ""),
                )

            except Exception as e:
                print(f"  Error: {str(e)}")
                traceback.print_exc()
                self._log_error(f"[process_single_profile] {url} :: {e}\n{traceback.format_exc()}")
                results.append({
                    'url': url,
                    'status': 'failed',
                    'error': str(e),
                    'profile_id': None
                })
                failed += 1
                self._set_excel_ignore_status(
                    excel_tracker,
                    excel_row_number=excel_row_number,
                    ignored=False,
                    reason="",
                    status="failed",
                    scrape_issue=str(e),
                )

            # Small delay to avoid overwhelming the system
            await asyncio.sleep(0.5)

        if self.incremental_sync_enabled and pending_sync_profiles:
            self._flush_incremental_sync_batch(
                pending_sync_profiles,
                is_final_flush=True,
            )

        self._save_excel_ignore_tracker(excel_tracker, excel_path)
        
        # Summary
        processed_total = total - ignored
        success_rate = (successful / processed_total * 100) if processed_total > 0 else 0.0
        sync_stats = self._get_incremental_sync_stats_snapshot()
        print(f"\n{'='*80}")
        print(f"[SUMMARY] Pipeline Complete!")
        print(f"  Total: {total}")
        print(f"  Successful: {successful}")
        print(f"  Ignored: {ignored}")
        print(f"  Failed: {failed}")
        print(f"  Success Rate (non-ignored): {success_rate:.1f}%")
        if self.incremental_sync_enabled:
            print(
                "  Sync: "
                f"batches={sync_stats.get('batches_completed', 0)}/"
                f"{sync_stats.get('batches_attempted', 0)}, "
                f"vectors_uploaded={sync_stats.get('vectors_uploaded', 0)}, "
                f"mongo_synced={sync_stats.get('profiles_synced_mongo', 0)}, "
                f"mongo_failed={sync_stats.get('profiles_failed_mongo', 0)}"
            )
        print(f"{'='*80}")

        registry_snapshot = self.source_registry.snapshot()

        return {
            'total': total,
            'successful': successful,
            'ignored': ignored,
            'failed': failed,
            'results': results,
            'output_dir': str(self.output_dir),
            'chunking_output_dir': str(self.chunking_output_dir),
            'source_registry': self.source_registry_path,
            'source_registry_snapshot': str(registry_snapshot) if registry_snapshot else None,
            'incremental_sync': sync_stats,
        }

    async def run_from_urls(
        self,
        urls: List[str],
        profile_name: Optional[str] = None,
        profile_url: Optional[str] = None,
        profile_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Run pipeline for a single profile using a list of seed URLs.
        """
        print("="*80)
        print("UNIFIED PIPELINE: URL LIST → Scraping → Cleaning → Chunking → JSON")
        print("="*80)

        if not urls:
            raise ValueError("No URLs provided")

        await self.initialize_scraper()

        result = await self.process_profile_from_urls(
            urls=urls,
            profile_name=profile_name,
            profile_url=profile_url,
            profile_id=profile_id,
        )

        status = result.get("status")
        successful = 1 if status == "success" else 0
        ignored = 1 if status == "ignored" else 0
        failed = 0 if status in ("success", "ignored") else 1
        if self.incremental_sync_enabled and status == "success":
            self._flush_incremental_sync_batch(
                [
                    {
                        "profile_id": result.get("profile_id"),
                        "profile_name": result.get("profile_name", "Unknown"),
                        "chunks_path": result.get("chunks_path"),
                    }
                ],
                is_final_flush=True,
            )
        sync_stats = self._get_incremental_sync_stats_snapshot()

        print(f"\n{'='*80}")
        print(f"[SUMMARY] Pipeline Complete!")
        print(f"  Total: 1")
        print(f"  Successful: {successful}")
        print(f"  Ignored: {ignored}")
        print(f"  Failed: {failed}")
        if self.incremental_sync_enabled:
            print(
                "  Sync: "
                f"batches={sync_stats.get('batches_completed', 0)}/"
                f"{sync_stats.get('batches_attempted', 0)}, "
                f"vectors_uploaded={sync_stats.get('vectors_uploaded', 0)}, "
                f"mongo_synced={sync_stats.get('profiles_synced_mongo', 0)}, "
                f"mongo_failed={sync_stats.get('profiles_failed_mongo', 0)}"
            )
        if "seed_urls_total" in result:
            print(f"  Seed URLs: {result.get('seed_urls_total', 0)}")
            print(f"  Seed URL Success: {result.get('seed_urls_successful', 0)}")
            print(f"  Seed URL Failed: {result.get('seed_urls_failed', 0)}")
        print(f"{'='*80}")

        registry_snapshot = self.source_registry.snapshot()

        return {
            "total": 1,
            "successful": successful,
            "ignored": ignored,
            "failed": failed,
            "results": [result],
            "output_dir": str(self.output_dir),
            "chunking_output_dir": str(self.chunking_output_dir),
            "source_registry": self.source_registry_path,
            "source_registry_snapshot": str(registry_snapshot) if registry_snapshot else None,
            "incremental_sync": sync_stats,
        }


async def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Unified Pipeline: Process Excel file to chunked JSON files'
    )
    parser.add_argument(
        'excel_path',
        type=str,
        nargs='?',
        default='profile.xlsx',
        help='Path to Excel file with profile URLs (default: profile.xlsx)'
    )
    parser.add_argument(
        '--urls-file',
        type=str,
        default=None,
        help='Path to a .txt file with one URL per line (treat as a single profile)'
    )
    parser.add_argument(
        '--profile-name',
        type=str,
        default=None,
        help='Optional profile name to use in URLs-file mode'
    )
    parser.add_argument(
        '--profile-url',
        type=str,
        default=None,
        help='Primary profile URL to store in output (defaults to first URL in file)'
    )
    parser.add_argument(
        '--output-dir',
        type=str,
        default=None,
        help='Output directory (default: output, or output/url_list_runs/<timestamp> in URLs-file mode)'
    )
    parser.add_argument(
        '--chunking-output-dir',
        type=str,
        default=None,
        help='Chunking output directory (default: <output_dir>/chunked_profiles)'
    )
    parser.add_argument(
        '--no-llm-chunking',
        action='store_true',
        help='Disable LLM-based section-aware chunking (faster but less accurate)'
    )
    parser.add_argument(
        '--llm-provider',
        type=str,
        default='openai',
        choices=['ollama', 'openai'],
        help='LLM provider for chunking (default: openai)'
    )
    parser.add_argument(
        '--llm-model',
        type=str,
        default='gpt-4o-mini',
        help='LLM model name (forced to gpt-4o-mini)'
    )
    parser.add_argument(
        '--llm-cleaning',
        action='store_true',
        help='Enable LLM-based text cleaning before section chunking'
    )
    parser.add_argument(
        '--no-llm-cleaning',
        action='store_true',
        help='Disable LLM-based text cleaning even when auto-enabled'
    )
    parser.add_argument(
        '--source-registry',
        type=str,
        default=None,
        help='Path to source registry JSONL (default: <output_dir>/source_registry.jsonl)'
    )
    parser.add_argument(
        '--strict-source-policy',
        action='store_true',
        default=None,
        help='Block sources with unknown license/robots disallow/paywall'
    )
    parser.add_argument(
        '--no-strict-source-policy',
        action='store_true',
        default=False,
        help='Disable strict source blocking even if env enables it'
    )
    parser.add_argument(
        '--search-query',
        type=str,
        default=None,
        help='Optional search query used to obtain URLs'
    )
    parser.add_argument(
        '--search-engine',
        type=str,
        default=None,
        help='Optional search engine used to obtain URLs'
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=None,
        help='Limit number of profiles to process'
    )
    parser.add_argument(
        '--start-from',
        type=int,
        default=0,
        help='Start processing from this index (for resuming)'
    )
    parser.add_argument(
        '--incremental-sync',
        action='store_true',
        help='Enable batchwise incremental Pinecone/Mongo sync during run'
    )
    parser.add_argument(
        '--no-incremental-sync',
        action='store_true',
        help='Disable incremental Pinecone/Mongo sync during run'
    )
    parser.add_argument(
        '--sync-batch-size',
        type=int,
        default=100,
        help='Number of successful profiles per incremental sync batch (default: 100)'
    )
    parser.add_argument(
        '--sync-pinecone-batch-size',
        type=int,
        default=50,
        help='Chunk embedding/upload batch size per incremental Pinecone flush (default: 50)'
    )
    parser.add_argument(
        '--skip-sync-pinecone',
        action='store_true',
        help='When incremental sync is enabled, skip Pinecone upload stage'
    )
    parser.add_argument(
        '--skip-sync-mongodb',
        action='store_true',
        help='When incremental sync is enabled, skip MongoDB sync stage'
    )
    parser.add_argument(
        '--skip-sync-indexes',
        action='store_true',
        help='When incremental sync is enabled, skip MongoDB index creation'
    )
    
    args = parser.parse_args()

    # Resolve output directories
    is_urls_mode = bool(args.urls_file)
    if args.output_dir:
        resolved_output_dir = args.output_dir
    else:
        if is_urls_mode:
            run_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            resolved_output_dir = os.path.join("output", "url_list_runs", run_stamp)
        else:
            resolved_output_dir = "output"

    if args.chunking_output_dir:
        resolved_chunking_output_dir = args.chunking_output_dir
    else:
        resolved_chunking_output_dir = os.path.join(resolved_output_dir, "chunked_profiles")
    
    # Fix Windows event loop if needed
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    # Create pipeline
    search_meta = {}
    if args.search_query:
        search_meta["query"] = args.search_query
    if args.search_engine:
        search_meta["engine"] = args.search_engine

    if args.no_strict_source_policy:
        strict_policy = False
    elif args.strict_source_policy is True:
        strict_policy = True
    else:
        strict_policy = None

    llm_cleaning_override: Optional[bool] = None
    if args.no_llm_cleaning:
        llm_cleaning_override = False
    elif args.llm_cleaning:
        llm_cleaning_override = True
    incremental_sync_override: Optional[bool] = None
    if args.no_incremental_sync:
        incremental_sync_override = False
    elif args.incremental_sync:
        incremental_sync_override = True

    pipeline = UnifiedPipeline(
        output_dir=resolved_output_dir,
        chunking_output_dir=resolved_chunking_output_dir,
        use_llm_chunking=not args.no_llm_chunking,
        use_llm_cleaning=llm_cleaning_override,
        llm_provider=args.llm_provider,
        llm_model="gpt-4o-mini",
        source_registry_path=args.source_registry,
        strict_source_policy=strict_policy,
        default_search_meta=search_meta,
        incremental_sync_enabled=incremental_sync_override,
        incremental_sync_batch_size=args.sync_batch_size,
        incremental_pinecone_batch_size=args.sync_pinecone_batch_size,
        incremental_skip_pinecone=args.skip_sync_pinecone,
        incremental_skip_mongo=args.skip_sync_mongodb,
        incremental_skip_indexes=args.skip_sync_indexes,
    )

    # Run pipeline (Excel mode or URLs-file mode)
    if args.urls_file:
        urls = load_urls_file(args.urls_file)
        summary = await pipeline.run_from_urls(
            urls=urls,
            profile_name=args.profile_name,
            profile_url=args.profile_url
        )
    else:
        summary = await pipeline.run(
            excel_path=args.excel_path,
            limit=args.limit,
            start_from=args.start_from
        )
    
    print("\nPipeline completed!")
    print(f"Check output in: {summary['output_dir']}")
    if summary['successful'] > 0:
        print(f"Chunked profiles in: {summary['chunking_output_dir']}")


if __name__ == "__main__":
    asyncio.run(main())
